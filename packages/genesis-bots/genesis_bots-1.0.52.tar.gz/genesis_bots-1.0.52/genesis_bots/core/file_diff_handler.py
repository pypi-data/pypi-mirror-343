import os
import time
import yaml
from git import Repo
from datetime import datetime
from difflib import unified_diff
from typing import List, Dict, Optional, Union
import pandas as pd
from openpyxl import load_workbook
import uuid
from genesis_bots.core.tools.image_tools import _image_analysis


class GitFileManager:

    @classmethod
    def get_default_git_repo_path(cls):
        return os.path.join(os.getcwd(), 'bot_git')


    def __init__(self, repo_path: str = None):
        """Initialize GitFileManager with a repository path"""

        self.repo_path = os.getenv('GIT_PATH', self.get_default_git_repo_path())

        try:
            # Create directory if it doesn't exist
            os.makedirs(self.repo_path, exist_ok=True)

            # Try to initialize repository
            try:
                self.repo = Repo(self.repo_path)
            except:
                # If repository doesn't exist, initialize it
                self.repo = Repo.init(self.repo_path)

                # Configure git user for initial commit
                with self.repo.config_writer() as git_config:
                    git_config.set_value('user', 'email', 'bot@example.com')
                    git_config.set_value('user', 'name', 'Bot')

                # Create an initial empty commit
                # First create an empty file to commit
                readme_path = os.path.join(self.repo_path, 'README.md')
                with open(readme_path, 'w') as f:
                    f.write('# Git Repository\nInitialized by Bot')

                # Stage and commit
                self.repo.index.add(['README.md'])
                self.repo.index.commit("Initial commit")

        except Exception as e:
            raise Exception(f"Failed to initialize git repository: {str(e)}")

    def list_files(self, path: str = None) -> Dict:
        """List all tracked files in the repository or specific path
        
        For large directories (>50 files total):
        - Lists only immediate files/directories in current path
        - Shows count of files under each subdirectory
        
        For very large directories (>100 files in current path):
        - Shows only first 100 files
        - Includes total count of remaining files
        """
        # Get all tracked files
        all_files = [str(item[0]) for item in self.repo.index.entries.keys()]
        
        # Filter by path if provided
        if path:
            all_files = [f for f in all_files if f.startswith(path)]
            base_path = path
        else:
            base_path = ""

        # Get immediate files/dirs in current path
        current_level_files = []
        subdirs = {}
        
        for file in all_files:
            rel_path = file[len(base_path):].lstrip('/')
            parts = rel_path.split('/')
            
            if len(parts) == 1:  # File in current directory
                current_level_files.append(file)
            else:  # File in subdirectory
                subdir = parts[0]
                if subdir not in subdirs:
                    subdirs[subdir] = 0
                subdirs[subdir] += 1

        # Prepare result
        result = {
            "success": True,
            "total_files": len(all_files),
            "base_path": base_path or "root",
            "full_path": os.path.join(self.repo_path, base_path) if base_path else self.repo_path
        }

        # Handle large directories
        if len(all_files) > 50:
            result["files"] = current_level_files
            result["subdirectories"] = {
                dir_name: f"{count} files" for dir_name, count in subdirs.items()
            }
            result["note"] = f"Only showing immediate files/directories in {result['base_path']} due to large file count"
        else:
            result["files"] = all_files

        # Handle very large current directories
        if len(current_level_files) > 100:
            remaining = len(current_level_files) - 100
            result["files"] = current_level_files[:100]
            result["note"] = f"Showing first 100 files in {result['base_path']}. {remaining} more files exist"

        return result

    def read_file(self, file_path: str) -> str:
        """Read contents of a file from the repository"""
        
        # Remove leading slash from file_path if repo_path doesn't start with slash
        if file_path.startswith('/') and not self.repo_path.startswith('/'):
            file_path = file_path.lstrip('/')

        if 'BOT_GIT:' in file_path:
            file_path = file_path.replace('BOT_GIT:', '')

        if file_path.startswith(self.repo_path):
            full_path = file_path
        else:
            full_path = os.path.join(self.repo_path, file_path)

        if not os.path.exists(full_path):
            raise FileNotFoundError(f"File not found: {file_path}")

        with open(full_path, 'r') as f:
            return f.read()

    def write_file(self, file_path: str, content: str, commit_message: str = None, **adtl_info) -> Dict:
        """Write content to a file and optionally commit changes"""
        try:
            # Ensure directory exists
            os.makedirs(os.path.dirname(os.path.join(self.repo_path, file_path)), exist_ok=True)

            # Only decode base64 if explicitly told to do so
            is_base64 = adtl_info.get('is_base64', False)
            if is_base64:
                try:
                    import base64
                    decoded_content = base64.b64decode(content)
                    # Write binary content
                    with open(os.path.join(self.repo_path, file_path), 'wb') as f:
                        f.write(decoded_content)
                except Exception as e:
                    return {"success": False, "error": f"Failed to decode base64 content: {str(e)}"}
            else:
                # Write text content
                with open(os.path.join(self.repo_path, file_path), 'w') as f:
                    f.write(content)

            # Add to git
            self.repo.index.add([file_path])

            # Prepare response
            result = {
                "success": True,
                "reminder": "If this file is related to any project, remember to record it as a project asset using _manage_project_assets"
            }

            # Add commit-specific message if applicable
            if commit_message:
                self.commit_changes(commit_message)
                result["message"] = "File written and changes committed successfully"
            else:
                result["message"] = "File written successfully. Remember to commit changes with git_action('commit', message='your message')"

            return result
        except Exception as e:
            return {"success": False, "error": str(e)}

    def generate_diff(self, old_content: str, new_content: str, context_lines: int = 3) -> str:
        """Generate a unified diff between two content strings"""
        old_lines = old_content.splitlines(keepends=True)
        new_lines = new_content.splitlines(keepends=True)

        return ''.join(unified_diff(
            old_lines, new_lines,
            fromfile='old',
            tofile='new',
            n=context_lines
        ))

    def validate_diff_format(self, diff_content: str) -> bool:
        """Validate that the provided content follows unified diff format"""
        try:
            lines = diff_content.splitlines()
            if not lines:
                return False

            # Basic unified diff format validation
            # Should start with --- and +++ lines
            has_header = False
            has_changes = False

            for line in lines:
                if line.startswith('--- '):
                    has_header = True
                    continue
                if line.startswith('+++ '):
                    has_header = True
                    continue
                if line.startswith('@@ '):
                    has_changes = True
                    continue
                if line.startswith('+') or line.startswith('-'):
                    has_changes = True

            return has_header and has_changes
        except:
            return False

    def apply_diff(self, file_path: str, diff_content: str, commit_message: str = None) -> Dict:
        """Apply a diff to a file"""
        try:
            from difflib import unified_diff, restore

            full_path = os.path.join(self.repo_path, file_path)

            # Read the current file content
            with open(full_path, 'r') as f:
                current_content = f.read()

            # Split content into lines
            current_lines = current_content.splitlines(True)

            # Parse and apply the diff
            patch_lines = diff_content.splitlines(True)
            new_lines = list(restore(patch_lines, 1))  # 1 means to apply the changes
            new_content = ''.join(new_lines)

            # Write the new content
            with open(full_path, 'w') as f:
                f.write(new_content)

            # Add and commit if requested
            self.repo.index.add([file_path])
            if commit_message:
                self.commit_changes(commit_message)
                return {
                    "success": True,
                    "message": "Diff applied and changes committed successfully",
                    "content": new_content
                }
            else:
                return {
                    "success": True,
                    "message": "Diff applied successfully. Remember to commit changes with git_action('commit', message='your message')",
                    "content": new_content
                }
        except Exception as e:
            return {"success": False, "error": str(e)}

    def commit_changes(self, message: str) -> Dict:
        """Commit staged changes to the repository"""
        try:
            self.repo.index.commit(message)
            return {"success": True, "message": "Changes committed successfully"}
        except Exception as e:
            return {"success": False, "error": str(e)}

    def get_commit_history(self, file_path: str = None, max_count: int = 10) -> List[Dict]:
        """Get commit history for the repository or specific file"""
        try:
            if file_path:
                commits = list(self.repo.iter_commits(paths=file_path, max_count=max_count))
            else:
                commits = list(self.repo.iter_commits(max_count=max_count))

            return [{
                "hash": str(commit.hexsha),
                "message": commit.message,
                "author": str(commit.author),
                "date": datetime.fromtimestamp(commit.committed_date),
                "files": list(commit.stats.files.keys())
            } for commit in commits]
        except Exception as e:
            return []

    def create_branch(self, branch_name: str) -> Dict:
        """Create a new branch"""
        try:
            new_branch = self.repo.create_head(branch_name)
            new_branch.checkout()
            return {"success": True, "message": f"Branch '{branch_name}' created and checked out"}
        except Exception as e:
            return {"success": False, "error": str(e)}

    def switch_branch(self, branch_name: str) -> Dict:
        """Switch to an existing branch"""
        try:
            self.repo.heads[branch_name].checkout()
            return {"success": True, "message": f"Switched to branch '{branch_name}'"}
        except Exception as e:
            return {"success": False, "error": str(e)}

    def get_current_branch(self) -> str:
        """Get the name of the current branch"""
        return self.repo.active_branch.name

    def get_file_status(self, file_path: str = None) -> Dict:
        """Get the status of files in the repository"""
        try:
            if file_path:
                status = self.repo.git.status('--porcelain', file_path)
            else:
                status = self.repo.git.status('--porcelain')

            modified = []
            untracked = []
            staged = []

            for line in status.splitlines():
                if line:
                    status_code = line[:2]
                    path = line[3:]
                    if status_code == '??':
                        untracked.append(path)
                    elif status_code == ' M' or status_code == 'M ':
                        modified.append(path)
                    elif status_code == 'A ':
                        staged.append(path)

            result = {
                "success": True,
                "modified": modified,
                "untracked": untracked,
                "staged": staged
            }

            # Add reminder if there are uncommitted changes
            if modified or untracked or staged:
                result["message"] = "There are uncommitted changes. Remember to commit them with git_action('commit', message='your message')"

            return result
        except Exception as e:
            return {"success": False, "error": str(e)}

    def remove_file(self, file_path: str, commit_message: str = None) -> Dict:
        """Remove a file from the repository"""
        try:
            full_path = os.path.join(self.repo_path, file_path)
            if not os.path.exists(full_path):
                return {"success": False, "error": f"File not found: {file_path}"}

            # Remove file from filesystem
            os.remove(full_path)

            # Remove from git
            self.repo.index.remove([file_path])

            # Prepare response
            result = {
                "success": True,
                "message": f"File {file_path} removed successfully"
            }

            # Commit if message provided
            if commit_message:
                self.commit_changes(commit_message)
                result["message"] += " and changes committed"

            return result
        except Exception as e:
            return {"success": False, "error": str(e)}

    def git_action(self, action: str, **kwargs) -> Dict:
        """
        Unified interface for all git operations.
        
        Actions:
            - list_files: List all tracked files (optional: path)
            - read_file: Read file contents (requires: file_path)
            - write_file: Write content to file (requires: file_path, content; optional: commit_message)
            - generate_diff: Generate diff between contents (requires: old_content, new_content; optional: context_lines)
            - apply_diff: Apply a unified diff to a file (requires: file_path, diff_content; optional: commit_message)
            - commit: Commit changes (requires: message)
            - get_history: Get commit history (optional: file_path, max_count)
            - create_branch: Create new branch (requires: branch_name)
            - switch_branch: Switch to branch (requires: branch_name)
            - get_branch: Get current branch name
            - get_status: Get file status (optional: file_path)
            - remove_file: Remove a file from the repository (requires: file_path; optional: commit_message)
        
        Returns:
            Dict containing operation result and any relevant data
        """

        try:
            action = action.lower()


            if action == "list_files":
                path = kwargs.get("path")
                if path and path.startswith('/'):
                    path = path[1:]

                files = self.list_files(path)
                return {"success": True, "files": files, "git_base_path_on_server_disk": self.repo_path}

            elif action == "read_file":
                if "file_path" not in kwargs:
                    return {"success": False, "error": "file_path is required"}
                # Check if file_path starts with / and return error if it does
                if kwargs["file_path"].startswith('/'):
                    return {"success": False, "error": "Please provide a relative file path without leading /"}
                
                file_path = os.path.join(self.repo_path, kwargs["file_path"])
                if file_path.lower().endswith(('.xls', '.xlsx')):
                    try:
                        # First read the data as before
                        if 'sheet_name' in kwargs:
                            df = pd.read_excel(file_path, sheet_name=kwargs['sheet_name'])
                        else:
                            df = pd.read_excel(file_path, sheet_name=None)

                        # Now extract images using openpyxl
                        wb = load_workbook(file_path)
                        image_paths = {}
                        
                        for sheet_name in wb.sheetnames:
                            sheet = wb[sheet_name]
                            sheet_images = []
                            
                            # Extract images from the sheet
                            for image in sheet._images:
                                # Generate unique filename for each image
                                image_ext = image.ref.split('.')[-1] if '.' in image.ref else 'png'
                                image_filename = f"{uuid.uuid4()}.{image_ext}"
                                
                                # Create the directory path for this thread
                                if 'thread_id' not in kwargs:
                                    return {"success": False, "error": "thread_id is required for saving images"}
                                    
                                save_dir = os.path.join('./runtime/downloaded_files', str(kwargs['thread_id']))
                                os.makedirs(save_dir, exist_ok=True)
                                
                                # Full path for the image
                                image_path = os.path.join(save_dir, image_filename)
                                
                                # Save image
                                with open(image_path, 'wb') as img_file:
                                    img_file.write(image._data())
                                
                                sheet_images.append({
                                    'path': image_path,
                                    'filename': image_filename,
                                    'position': f"Row: {image.anchor._from.row}, Col: {image.anchor._from.col}"
                                })
                            
                            if sheet_images:
                                image_paths[sheet_name] = sheet_images

                        # Prepare the content response with analyzed images
                        total_image_count = 0

                        if isinstance(df, dict):
                            # Multiple sheets - convert each sheet to JSON
                            content = {}
                            for sheet_name, sheet_df in df.items():
                                sheet_images = image_paths.get(sheet_name, [])
                                # Analyze images for this sheet
                                for image_info in sheet_images:
                                    if total_image_count < 20:
                                        analysis_result = _image_analysis(
                                            query="Describe in detail what you see in this image, including the full text of any text in the image",
                                            file_name=image_info['filename'],
                                            thread_id=kwargs['thread_id']
                                        )
                                        
                                        if analysis_result.get('success'):
                                            image_info['analysis'] = analysis_result['data']
                                        else:
                                            image_info['analysis'] = "Failed to analyze image"
                                    elif total_image_count == 20:
                                        image_info['analysis'] = "Only the first 20 images were automatically analyzed. You can use the _image_analysis() function on this file if you need to analyze additional images."
                                    # For images after 21, we don't add an analysis field at all
                                    
                                    total_image_count += 1

                                content[sheet_name] = {
                                    'data': sheet_df.to_json(orient='records'),
                                    'images': sheet_images
                                }
                        else:
                            # Single sheet
                            sheet_images = image_paths.get(list(wb.sheetnames)[0], [])
                            # Analyze images
                            for image_info in sheet_images:
                                if total_image_count < 20:
                                    analysis_result = _image_analysis(
                                        query="Describe what you see in this image",
                                        file_name=image_info['filename'],
                                        thread_id=kwargs['thread_id']
                                    )
                                    
                                    if analysis_result.get('success'):
                                        image_info['analysis'] = analysis_result['data']
                                    else:
                                        image_info['analysis'] = "Failed to analyze image"
                                elif total_image_count == 20:
                                    image_info['analysis'] = "Only the first 20 images were automatically analyzed. You can use analyze_image() on this file if you need to analyze additional images."
                                # For images after 21, we don't add an analysis field at all
                                
                                total_image_count += 1

                            content = {
                                'data': df.to_json(orient='records'),
                                'images': sheet_images
                            }

                        return {
                            "success": True, 
                            "content": content, 
                            "is_excel": True,
                            "has_images": bool(image_paths),
                            "note": "Images were found in this Excel file. The first 20 images have been automatically analyzed. You can use display_image() to show specific images to the user. Use the image paths from the response with these functions."
                        }
                    except Exception as e:
                        return {"success": False, "error": f"Failed to read Excel file: {str(e)}"}
                else:
                    # For other files, use normal read
                    content = self.read_file(file_path)
                    return {"success": True, "content": content}

            elif action == "write_file":
                if "file_content" in kwargs and "content" not in kwargs:
                    kwargs["content"] = kwargs["file_content"]
                if "content" not in kwargs and "new_content" in kwargs:
                    kwargs["content"] = kwargs["new_content"]
                if "file_path" not in kwargs or "content" not in kwargs:
                    return {"success": False, "error": "file_path and content are required"}
                # Check if file_path starts with / and return error if it does
                if kwargs["file_path"].startswith('/'):
                    return {"success": False, "error": "Please provide a relative file path without leading /"}
                
                # Extract the standard parameters
                file_path = kwargs["file_path"]
                content = kwargs["content"]
                commit_message = kwargs.get("commit_message")
                
                # Pass through any additional parameters that aren't standard
                adtl_info = {k: v for k, v in kwargs.items() 
                             if k not in ["file_path", "content", "commit_message", 
                                         "file_content", "new_content"]}
                
                # Pass through any additional parameters
                return self.write_file(
                    file_path,
                    content,
                    commit_message,
                    **adtl_info  # Pass any additional parameters
                )

            elif action == "generate_diff":
                if "old_content" not in kwargs or "new_content" not in kwargs:
                    return {"success": False, "error": "old_content and new_content are required"}
                diff = self.generate_diff(
                    kwargs["old_content"],
                    kwargs["new_content"],
                    kwargs.get("context_lines", 3)
                )
                return {"success": True, "diff": diff}

            elif action == "apply_diff":
                if "file_path" not in kwargs or "diff_content" not in kwargs:
                    return {
                        "success": False,
                        "error": "file_path and diff_content are required"
                    }
                return self.apply_diff(
                    kwargs["file_path"],
                    kwargs["diff_content"],
                    kwargs.get("commit_message")
                )

            elif action == "commit":
                if "message" not in kwargs:
                    return {"success": False, "error": "commit message is required"}
                return self.commit_changes(kwargs["message"])

            elif action == "get_history":
                history = self.get_commit_history(
                    kwargs.get("file_path"),
                    kwargs.get("max_count", 10)
                )
                return {"success": True, "history": history}

            elif action == "create_branch":
                if "branch_name" not in kwargs:
                    return {"success": False, "error": "branch_name is required"}
                return self.create_branch(kwargs["branch_name"])

            elif action == "switch_branch":
                if "branch_name" not in kwargs:
                    return {"success": False, "error": "branch_name is required"}
                return self.switch_branch(kwargs["branch_name"])

            elif action == "get_branch":
                branch = self.get_current_branch()
                return {"success": True, "branch": branch}

            elif action == "get_status":
                status = self.get_file_status(kwargs.get("file_path"))
                return status

            elif action == "remove_file":
                if "file_path" not in kwargs:
                    return {"success": False, "error": "file_path is required"}
                return self.remove_file(
                    kwargs["file_path"],
                    kwargs.get("commit_message")
                )

            else:
                return {"success": False, "error": f"Unknown action: {action}"}

        except Exception as e:
            return {"success": False, "error": str(e)}
