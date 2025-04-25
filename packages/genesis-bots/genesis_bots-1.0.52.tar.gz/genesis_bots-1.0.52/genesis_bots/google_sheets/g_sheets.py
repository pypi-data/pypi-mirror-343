# Commented out section is for using OAuth instead of creating a service account

import os.path

from google.oauth2.credentials import Credentials as Creds_Oauth
from google.oauth2.service_account import Credentials as Creds_Service
from googleapiclient.discovery import build
from googleapiclient.errors import HttpError
from googleapiclient.http import MediaIoBaseDownload, MediaFileUpload
from datetime import datetime

# import mimetypes
import os, json

from genesis_bots.google_sheets.format_g_sheets import format_genesis_g_sheets

## test
from concurrent.futures import ThreadPoolExecutor
import time
from tenacity import retry, stop_after_attempt, wait_exponential

import openpyxl
import requests
from io import BytesIO

from genesis_bots.core.logging_config import logger
import re

# If modifying these scopes, delete the file token.json.
SCOPES = [
    "https://www.googleapis.com/auth/drive.file",
    "https://www.googleapis.com/auth/documents",
    "https://www.googleapis.com/auth/drive",
    "https://www.googleapis.com/auth/spreadsheets"
]

_g_creds = None
root_folder = None

def load_creds():
    root_folder = get_root_folder_id()
    if os.path.exists('g-workspace-oauth-credentials.json'):
        return get_g_creds_oauth()
    else:
        return get_g_creds_service_account()

def get_g_creds_oauth():
    global _g_creds, root_folder
    if _g_creds is None:
        OAUTH_KEY_FILE = f"g-workspace-oauth-credentials.json"
        if not os.path.exists(OAUTH_KEY_FILE):
            logger.info(f"Authorized user file not found: {OAUTH_KEY_FILE}")
        try:
            _g_creds = Creds_Oauth.from_authorized_user_file(OAUTH_KEY_FILE, SCOPES)
            json_creds = json.loads(creds.to_json())
            root_folder = 'root'
            logger.info(f"Creds_Oauth loaded: {json_creds}")
        except Exception as e:
            logger.error(f"Error loading credentials: {e}")
            _g_creds = None
            return False
    return _g_creds

def get_g_creds_service_account():
    global _g_creds, root_folder
    SERVICE_ACCOUNT_FILE = f"g-workspace-sa-credentials.json"
    try:
        # Authenticate using the service account JSON file
        creds = Creds_Service.from_service_account_file(
            SERVICE_ACCOUNT_FILE, scopes=SCOPES
        )
        root_folder = get_root_folder_id()
    except Exception as e:
        print(f"Error loading credentials: {e}")
        return None

    return creds

def use_service_account():
    global root_folder
    file_path = "g-workspace-credentials.json"
    if os.path.exists(file_path):
        os.remove(file_path)
        root_folder = get_root_folder_id()
        logger.info(f"Deleted oauth credentials {file_path}.  Google Drive will now use service account.")
    else:
        logger.info(f"Oath file not found: {file_path}")


def column_to_number(letter: str) -> int:
    num = 0
    for char in letter:
        num = num * 26 + (ord(char.upper()) - ord('A') + 1)
    return num

def number_to_column(num: int) -> str:
    result = ""
    while num > 0:
        num -= 1
        result = chr(num % 26 + 65) + result
        num //= 26
    return result

def parse_cell_range(cell_range):
    cell_range = re.sub(r'[^A-Za-z0-9:]', '', cell_range)
    if cell_range.count(':') > 1:
        parts = cell_range.split(':')
        cell_range = parts[0] + ':' + parts[1]
    match = re.match(r"([A-Za-z]+)(\d+)(?::([A-Za-z]+)(\d+))?", cell_range)
    if not match:
        raise ValueError("Invalid cell range format")

    start_col, start_row, end_col, end_row = match.groups()
    start_col_num = column_to_number(start_col)
    start_row_num = int(start_row)

    if end_col and end_row:
        end_col_num = column_to_number(end_col)
        end_row_num = int(end_row)
    else:
        end_col_num = start_col_num
        end_row_num = start_row_num

    num_cells = (end_col_num - start_col_num + 1) * (end_row_num - start_row_num + 1)
    return start_col_num, start_row_num, end_col_num, end_row_num, num_cells

def read_g_doc(doc_id, creds=None):
    creds = load_creds()

    try:
        service = build("docs", "v1", credentials=creds)
        document = service.documents().get(documentId=doc_id).execute()
        content = document.get('body').get('content')

        service._http.http.close()

        text = ""
        for element in content:
            if 'paragraph' in element:
                for text_run in element.get('paragraph').get('elements'):
                    text += text_run.get('textRun').get('content', '')

        return {"Success": True, "Text": text}

    except HttpError as error:
        print(f"An error occurred: {error}")
        return None

def create_g_doc(data, g_doc_title='Untitled Document', folder_id=None, creds=None):
    logger.info('Entering create_g_doc')
    creds = load_creds()

    try:
        logger.info('Setting up services')
        docs_service = build("docs", "v1", credentials=creds)
        drive_service = build("drive", "v3", credentials=creds)

        logger.info(f'Creating doc {g_doc_title}...')
        body = {"title": g_doc_title}
        doc = docs_service.documents().create(body=body).execute()
        doc_id = doc.get("documentId")

        logger.info('Inserting text...')
        requests = [{"insertText": {"location": {"index": 1}, "text": data}}]
        docs_service.documents().batchUpdate(documentId=doc_id, body={"requests": requests}).execute()

        if folder_id:
            logger.info(f'Moving document to folder {folder_id}')
            drive_service.files().update(fileId=doc_id, addParents=folder_id).execute()

        return {"Success": True, "Document ID": doc_id}

    except HttpError as error:
        print(f"An error occurred: {error}")
        return {"Success": False, "Error": str(error)}

def append_g_doc(doc_id, data, creds=None):
    creds = load_creds()

    try:
        docs_service = build("docs", "v1", credentials=creds)

        # Get the current document content
        document = docs_service.documents().get(documentId=doc_id).execute()
        end_index = document.get('body').get('content')[-1].get('endIndex')

        # Prepare the request to append text
        requests = [{"insertText": {"location": {"index": end_index - 1}, "text": data}}]
        docs_service.documents().batchUpdate(documentId=doc_id, body={"requests": requests}).execute()

        return {"Success": True, "Document ID": doc_id}

    except HttpError as error:
        print(f"An error occurred: {error}")
        return {"Success": False, "Error": str(error)}

def update_g_doc(doc_id, data, creds=None):
    """
    Update a Google Doc's content by replacing all existing content with new data.

    Args:
        doc_id (str): The ID of the document to update
        data (str): The new content to write
        creds: Optional credentials object

    Returns:
        dict: Result containing Success status and Document ID or Error
    """
    creds = load_creds()

    try:
        docs_service = build("docs", "v1", credentials=creds)

        # First, get the current document to find its length
        document = docs_service.documents().get(documentId=doc_id).execute()
        end_index = document.get('body').get('content')[-1].get('endIndex', 1)

        # Prepare the requests to replace content
        requests = [
            # Delete all content if document is not empty
            {
                "deleteContentRange": {
                    "range": {
                        "startIndex": 1,
                        "endIndex": end_index - 1
                    }
                }
            } if end_index > 1 else None,
            # Insert new content
            {
                "insertText": {
                    "location": {"index": 1},
                    "text": data
                }
            }
        ]

        # Remove None entries
        requests = [r for r in requests if r is not None]

        # Execute the update
        docs_service.documents().batchUpdate(
            documentId=doc_id,
            body={"requests": requests}
        ).execute()

        return {
            "Success": True,
            "Document ID": doc_id
        }

    except HttpError as error:
        logger.error(f"HTTP error updating document: {error}")
        return {"Success": False, "Error": str(error)}
    except Exception as e:
        logger.error(f"Error updating document: {e}")
        return {"Success": False, "Error": str(e)}

def get_root_folder_id():
    global root_folder

    if root_folder:
        return root_folder

    from genesis_bots.connectors import get_global_db_connector
    db_adapter = get_global_db_connector()

    connection = db_adapter.connection
    cursor = connection.cursor()

    select_query = f"""
    SELECT value
    FROM {db_adapter.schema}.EXT_SERVICE_CONFIG
    WHERE parameter = 'shared_folder_id' AND ext_service_name = 'g-sheets'
    """
    cursor.execute(select_query)
    result = cursor.fetchone()

    cursor.close()

    if result:
        root_folder = result[0]
        return result[0]
    else:
        root_folder = None
        return None

def set_root_folder_id(db_adapter, folder_id):
    global root_folder
    root_folder = folder_id
    connection = db_adapter.connection
    cursor = connection.cursor()

    # Check if the key exists
    select_query = f"""
    SELECT COUNT(*)
    FROM {db_adapter.schema}.EXT_SERVICE_CONFIG
    WHERE parameter = 'shared_folder_id' AND ext_service_name = 'g-sheets'
    """
    cursor.execute(select_query)
    key_exists = cursor.fetchone()[0]

    if key_exists:
        # Update the existing key
        update_query = f"""
        UPDATE {db_adapter.schema}.EXT_SERVICE_CONFIG
        SET value = %s, updated = CURRENT_TIMESTAMP
        WHERE parameter = 'shared_folder_id' AND ext_service_name = 'g-sheets'
        """
        cursor.execute(update_query, (folder_id,))
    else:
        # Insert a new row
        insert_query = f"""
        INSERT INTO {db_adapter.schema}.EXT_SERVICE_CONFIG (parameter, value, ext_service_name, created, updated)
        VALUES ('shared_folder_id', %s, 'g-sheets', CURRENT_TIMESTAMP, CURRENT_TIMESTAMP)
        """
        cursor.execute(insert_query, (folder_id,))

    connection.commit()
    cursor.close()

def insert_into_g_drive_file_version_table(self, data):
    """
    Insert a new row into the G_DRIVE_FILE_VERSION table.

    Args:
        data (dict): A dictionary containing the following keys:
            - g_file_id
            - g_file_name
            - g_file_type
            - g_file_parent_id
            - g_file_size
            - g_file_version
    """
    # Assuming you have a database connection established
    connection = self.connection
    cursor = connection.cursor()

    insert_query = f"""
    INSERT INTO {self.schema}.G_DRIVE_FILE_VERSION (g_file_id, g_file_name, g_file_type, g_file_parent_id, g_file_size, g_file_version)
    VALUES (%s, %s, %s, %s, %s, %s)
    """
    cursor.execute(
        insert_query,
        (
            data["g_file_id"],
            data["g_file_name"],
            data["g_file_type"],
            data["g_file_parent_id"],
            data["g_file_size"],
            data["g_file_version"],
        ),
    )

    connection.commit()
    cursor.close()
    return {"Success": True, "Message": "File version inserted."}


def update_g_drive_file_version_table(
    db_adapter, g_file_id, g_file_version, g_file_name, g_file_size, g_folder_id, g_file_type
):
    """
    Update the version of a file in the G_DRIVE_FILE_VERSION table.

    Args:
        g_file_id (str): The ID of the file.
        g_file_version (str): The new version of the file.
        g_file_name (str): The name of the file.
        g_file_size (str): The size of the file.
        g_folder_id (str): The ID of the folder containing the file.
        g_file_type (str): The type of the file.
    """
    connection = db_adapter.connection
    cursor = connection.cursor()

    # Check if the file ID exists in the table
    select_query = f"""
    SELECT COUNT(*)
    FROM {db_adapter.schema}.G_DRIVE_FILE_VERSION
    WHERE g_file_id = %s
    """
    cursor.execute(select_query, (g_file_id,))
    file_exists = cursor.fetchone()[0]

    if file_exists == 0:
        insert_query = f"""
        INSERT INTO {db_adapter.schema}.G_DRIVE_FILE_VERSION (g_file_id, g_file_version, g_file_name, g_file_size, g_file_parent_id, g_file_type)
        VALUES (%s, %s, %s, %s, %s, %s)
        """
        cursor.execute(insert_query, (g_file_id, g_file_version, g_file_name, g_file_size, g_folder_id, g_file_type))
    else:
        update_query = f"""
        UPDATE {db_adapter.schema}.G_DRIVE_FILE_VERSION
        SET g_file_version = %s
        WHERE g_file_id = %s
        """
        cursor.execute(update_query, (g_file_version, g_file_id))

    connection.commit()
    cursor.close()

    return {"Success": True, "Message": "File version updated."}


def get_g_file_comments(file_id, user='Unknown User'):
    """
    Get comments on a Google Sheets document.

    Args:
        file_id (str): The ID of the file.

    Returns:
        list: A list of comments on the document.
    """
    creds = load_creds()

    try:
        service = build("drive", "v3", credentials=creds)

        # Get the comments on the document
        comments = (
            service.comments()
            .list(fileId=file_id, fields="comments(anchor, id,content,author(kind, displayName,emailAddress),replies(id,content,author(displayName,emailAddress),htmlContent))")
            .execute()
        )

        # Get the web link to the file
        file_metadata = service.files().get(fileId=file_id, fields="webViewLink").execute()
        file_url = file_metadata.get("webViewLink")

        flat_comments = []

        # Add the URL to each comment
        for comment in comments.get("comments", []):
            comment["url"] = f"{file_url}?comment={comment['id']}"
            for reply in comment.get("replies", []):
                reply["url"] = f"{file_url}?comment={comment['id']}&reply={reply['id']}"

            flat_comments.append(comment['content'])

        request = service.files().export_media(fileId=file_id, mimeType="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        fh = BytesIO()
        downloader = MediaIoBaseDownload(fh, request)
        done = False
        while done is False:
            status, done = downloader.next_chunk()
            logger.info("Download %d%%" % int(status.progress() * 100))
        fh.seek(0)
        workbook = openpyxl.load_workbook(filename=fh, data_only=False)
        worksheet = workbook['Sheet1']
        res = []
        for i, row in enumerate(worksheet.iter_rows()):
            for j, cell in enumerate(row):
                if cell.comment:
                    try:
                        comment_index = flat_comments.index(cell.comment.text.split("\n", 1)[0])
                    except ValueError:
                        continue
                    comments['comments'][comment_index]["cellRC"] = number_to_column(j + 1).strip() + str(i + 1).strip()
                    comments["comments"][comment_index]["columnIndex"] = (
                        number_to_column(j + 1)
                    )
        service._http.http.close()

        return comments.get("comments", [])

    except Exception as e:
        logger.error(f"An error occurred: {str(e)}")
        return None


def add_reply_to_g_file_comment(
    file_id=None, comment_id=None, reply_content=None, g_file_comment_id=None, creds=None, user='Unknown User'
):
    """
    Add a reply to a comment on a Google Drive file.

    Args:
        user (str): The user associated with the service account.
        file_id (str): The ID of the file.
        comment_id (str): The ID of the comment.
        reply_content (str): The content of the reply.

    Returns:
        dict: The created reply.
    """

    creds = load_creds()

    try:
        service = build("drive", "v3", credentials=creds)

        # Create the reply
        reply_body = {"content": reply_content}
        created_reply = (
            service.replies()
            .create(
                fileId=file_id,
                commentId=g_file_comment_id,
                body=reply_body,
                fields="id,content",
            )
            .execute()
        )

        service._http.http.close()

        logger.info(f"Reply added: {created_reply['content']}")
        return created_reply

    except Exception as e:
        logger.error(f"An error occurred: {str(e)}")
        return None

def get_g_file_web_link(file_id, creds=None):
    """
    Get the web link to a file in Google Drive.

    Args:
        file_id (str): The ID of the file.

    Returns:
        str: The web link to the file.
    """

    creds = load_creds()

    try:
        service = build("drive", "v3", credentials=creds)

        # Get the file metadata including the webViewLink
        file_metadata = service.files().get(fileId=file_id, fields="name, webViewLink, parents").execute()

        service._http.http.close()

        return {
            "Success": True,
            "Name": file_metadata.get("name"),
            "URL": file_metadata.get("webViewLink"),
            "Folder ID": (
                file_metadata.get("parents")[0]
                if file_metadata.get("parents")
                else None
            ),
        }

    except Exception as e:
        return {"Success": False, "Error": str(e)}

def find_g_file_by_name(file_name, creds=None):
    """
    Find all files in Google Drive by their name.

    Args:
        file_name (str): The name of the file.

    Returns:
        dict: A list of file metadata if found, otherwise None.
    """

    creds = load_creds()

    try:
        service = build("drive", "v3", credentials=creds)

        # Search for the files by name
        query = f"name='{file_name}'"
        response = service.files().list(q=query, fields="files(id, name, webViewLink, createdTime)").execute()
        files = response.get("files", [])

        service._http.http.close()

        if files:
            return {"Success": True, "Files": files}
        else:
            return {"Success": False, "Error": "Files not found"}

    except Exception as e:
        return {"Success": False, "Error": str(e)}

def get_g_folder_directory(folder_id=None, creds=None, db_adapter=None):
    """
    Get all files in a Google Drive folder.

    Args:
        folder_id (str): The ID of the folder.

    Returns:
        list: A list of files in the folder.
    """
    global root_folder
    logger.info(f"Entering get_g_folder_directory with folder_id: {folder_id}")

    creds = load_creds()

    if not folder_id:
        folder_id = root_folder

    try:
        service = build("drive", "v3", credentials=creds)

        # Get the list of files in the folder
        query = f"'{folder_id}' in parents and trashed = false"
        fields = "files(id, name, mimeType, createdTime, modifiedTime, webViewLink, size)"

        files = []
        page_token = None
        while True:
            try:
                response = service.files().list(
                    q=query,
                    fields=fields,
                    pageSize=1000,
                    orderBy="name",
                    pageToken=page_token,
                ).execute()

                files.extend(response.get("files", []))
                page_token = response.get("nextPageToken")

                if not page_token:
                    break

            except HttpError as error:
                logger.error(f"Error listing files: {error}")
                return {"Success": False, "Error": f"API error: {str(error)}"}

        logger.info(f"Found {len(files)} files in folder {folder_id}")

        # Format response
        file_list = [{
            "id": f.get("id"),
            "name": f.get("name"),
            "type": f.get("mimeType"),
            "created": f.get("createdTime"),
            "modified": f.get("modifiedTime"),
            "url": f.get("webViewLink"),
            "size": f.get("size")
        } for f in files]

        service._http.http.close()

        return {
            "Success": True,
            "Files": file_list,
            "Total": len(file_list)
        }

    except Exception as e:
        return {"Success": False, "Error": str(e)}

def add_g_file_comment(
    file_id=None,
    content=None,
    creds=None,
    user='Unknown User'
):
    """
    Add a comment to a Google Drive file.

    Args:
        user (str): The user associated with the service account.
        file_id (str): The ID of the file.
        content (str): The content of the comment.

    Returns:
        dict: The created comment.
    """

    creds = load_creds()

    try:
        service = build("drive", "v3", credentials=creds)

        # Create the comment
        body = {"content": content}
        created_comment = (
            service.comments()
            .create(fileId=file_id, body=body, fields="id,content")
            .execute()
        )

        service._http.http.close()

        logger.info(f"Comment added: {created_comment['content']}")
        return created_comment

    except Exception as e:
        logger.error(f"An error occurred: {str(e)}")
        return None


def get_g_folder_web_link(folder_id, creds):
    """
    Get the web link to a folder in Google Drive.

    Args:
        folder_id (str): The ID of the folder.

    Returns:
        str: The web link to the folder.
    """
    creds = load_creds()

    try:
        # Authenticate using the service account JSON file
        service = build("drive", "v3", credentials=creds)

        # Get the folder metadata including the webViewLink
        folder = service.files().get(fileId=folder_id, fields="id, name, webViewLink").execute()

        # logger.info the folder details
        logger.info(f"Folder ID: {folder.get('id')}")
        logger.info(f"Folder Name: {folder.get('name')}")
        logger.info(f"Web View Link: {folder.get('webViewLink')}")

        service._http.http.close()

        return folder.get("webViewLink")

    except Exception as e:
        logger.info(f"An error occurred: {str(e)}")
        return None


def get_g_file_version(g_file_id = None, creds = None, db_adapter = None):
    """
    Get the version number of a file in Google Drive.

    Args:
        file_id (str): The ID of the file.

    Returns:
        int: The version number of the file.
    """

    creds = load_creds()

    service = build("drive", "v3", credentials=creds)

    # Get the file metadata including the version number
    file_metadata = service.files().get(fileId=g_file_id, fields="version, name, size, parents").execute()

    version = file_metadata.get("version")
    file_name = file_metadata.get("name")
    file_size = file_metadata.get("size")
    parent_folder_id = file_metadata.get("parents")[0] if file_metadata.get("parents") else None
    g_file_type = 'sheet'

    service._http.http.close()

    update_g_drive_file_version_table(db_adapter, g_file_id, version, file_name, file_size, parent_folder_id, g_file_type)

    # logger.info the file version
    return version


# def upload_file_to_folder(path_to_file, parent_folder_id):
#     creds = Creds_Oauth.from_service_account_file(
#             SERVICE_ACCOUNT_FILE, scopes=SCOPES
#         )
#     service = build("drive", "v3", credentials=creds)

#     file_path = os.path(path_to_file)
#     filename = os.path.basename(file_path)
#     mime_type = mimetypes.guess_type(file_path)

#     file_metadata = {"name": filename}
#     if parent_folder_id:
#         file_metadata["parents"] = [parent_folder_id]

#     media = MediaFileUpload(file_path, mimetype=mime_type[0])
#     file = (
#         service.files().create(body=file_metadata, media_body=media, fields="id").execute()
#     )
#     logger.info(f'File ID: "{file.get("id")}".')
#     return file.get("id")


def process_row(args):
    self, row, stage_column_index, stage_column_folder_ids, creds = args
    row_values = list(row.values())

    for j, row_value in enumerate(row_values):
        if isinstance(row_value, datetime):
            row_values[j] = row_value.strftime("%Y-%m-%d %H:%M:%S")
        elif len(stage_column_index) > 0 and j in stage_column_index and row_value:
            if len(row_value) < 1 or not row_value.startswith('@'):
                continue

            parts = row_value.split(".")
            path = parts[2].split("/")
            stage = path[0]

            file_contents = self.read_file_from_stage(
                parts[0].replace('@',''),
                parts[1],
                stage,
                "/".join(path[1:]) + '.' + parts[-1],
                True,
            )

            filename = path[-1] + '.' + parts[-1]
            stage_folder_id = stage_column_folder_ids[stage_column_index.index(j)]

            webLink = save_text_to_google_file_with_retry(
                self, stage_folder_id, filename, file_contents, creds
            )

            # Remove any quotes around the URL and filename to prevent formula errors
            webLink = webLink.replace('"', '') if webLink else ''
            filename = filename.replace('"', '')
            row_values[j] = f'=HYPERLINK("{webLink}")'

    return row_values

@retry(
    stop=stop_after_attempt(3),
    wait=wait_exponential(multiplier=1, min=4, max=10),
    retry_error_callback=lambda retry_state: None
)
def save_text_to_google_file_with_retry(*args, **kwargs):
    return save_text_to_google_file(*args, **kwargs)

def save_text_to_google_file(
    self, shared_folder_id, file_name, text = "No text in file", creds=None
):
    if not text or isinstance(text, dict):
        text = "No text received in save_text_to_google_file."

    if not creds:
        OAUTH_KEY_FILE = f"g-workspace-sa-credentials.json"
        if not os.path.exists(OAUTH_KEY_FILE):
            logger.info(f"Service account file not found: {OAUTH_KEY_FILE}")
        try:
            creds = Creds_Oauth.from_authorized_user_file(OAUTH_KEY_FILE, SCOPES)

            logger.info(f"Creds_Oauth loaded: {creds}")
        except Exception as e:
            logger.error(f"Error loading credentials: {e}")
            return False

    docs_service = build("docs", "v1", credentials=creds)
    drive_service = build("drive", "v3", credentials=creds)

    # Check if a document with the same name already exists in the shared folder
    query = f"'{shared_folder_id}' in parents and name='{file_name}' and mimeType='application/vnd.google-apps.document'"
    response = (
        drive_service.files().list(q=query, fields="files(id, name)").execute()
    )
    files = response.get("files", [])

    if files:
        for file in files:
            logger.info(
                f"Deleting existing file: {file.get('name')} (ID: {file.get('id')})"
            )
            docs_service.files().delete(fileId=file.get("id")).execute()

    # Create a new document
    if not file_name:
        file_name = "genesis_" + datetime.now().strftime("%m%d%Y_%H:%M:%S")

    body = {"title": file_name}
    doc = docs_service.documents().create(body=body).execute()
    logger.info("Created document with title: {0}".format(doc.get("title")))
    doc_id = doc.get("documentId")
    logger.info(f"Document ID: {doc_id}")

    # Move the document to shared folder
    if shared_folder_id:
        file = (
            drive_service.files()
            .update(
                fileId=doc_id,
                addParents=shared_folder_id,
                fields="id, parents",
            )
            .execute()
        )
        logger.info(f"File moved to folder: {file} | Parent folder {file['parents'][0]}")

    # Verify the new document exists in Google Drive
    try:
        file_verify = (
            drive_service.files()
            .get(fileId=doc_id, fields="id, name, parents, webViewLink")
            .execute()
        )
        logger.info(f"File store confirmed: {file_verify}")
    except:
        raise Exception("Error creating document in Google Drive")

    parent = (
        drive_service.files().get(fileId=shared_folder_id, fields="id, name").execute()
    )
    logger.info(f"Parent folder name: {parent.get('name')} (ID: {parent.get('id')})")

    requests = [{"insertText": {"location": {"index": 1}, "text": text}}]

    result = (
        docs_service.documents()
        .batchUpdate(documentId=doc_id, body={"requests": requests})
        .execute()
    )

    logger.info("Document content updated: ", result)

    # Add to G_DRIVE_FILE_VERSION table
    g_file_version_data = {
        "g_file_id": doc_id,
        "g_file_name": file_name,
        "g_file_type": "application/vnd.google-apps.document",
        "g_file_parent_id": shared_folder_id,
        "g_file_size": str(len(text)),
        "g_file_version": "1"
    }

    insert_into_g_drive_file_version_table(self, g_file_version_data)

    return {
            "Success": True,
            "URL": file_verify.get("webViewLink"),
        }


def create_folder_in_folder(folder_name, parent_folder_id):
    creds = load_creds()

    service = build("drive", "v3", credentials=creds)

    file_metadata = {
        "name": folder_name,
        "parents": [parent_folder_id],
        "mimeType": "application/vnd.google-apps.folder",
    }

    file = service.files().create(body=file_metadata, fields="id").execute()

    service._http.http.close()

    logger.info(f'Folder ID: {file.get("id")} | Folder name: {folder_name}')

    return file.get("id")


def create_google_sheet_from_export(self, shared_folder_id, title, data):
    """
    Creates a Google Sheet with the given title and table data and moves it
    from the service account to the shared folder.
    Loads pre-authorized user credentials from the environment.
    """

    if not data:
        return {"Success": True, "message": "No data provided."}

    creds = load_creds()

    try:
        # service = build("sheets", "v4", credentials=creds)
        drive_service = build("drive", "v3", credentials=creds)

        service = build("sheets", "v4", credentials=creds)

        spreadsheet = {"properties": {"title": title}}
        spreadsheet = (
            service.spreadsheets()
            .create(body=spreadsheet, fields="spreadsheetId")
            .execute()
        )

        ss_id = spreadsheet.get("spreadsheetId")
        logger.info(f"Spreadsheet ID: {ss_id}")
        keys = list(data[0].keys())
        columns = [keys]

        # Check for stage links
        stage_column_index = [i for i, key in enumerate(keys) if key.endswith("_STAGE_LINK")]
        stage_column_folder_names = [key.replace("_STAGE_LINK", "") for key in keys if key.endswith("_STAGE_LINK")]
        stage_column_folder_ids = []

        # Create folder top level folder
        top_level_folder_id = create_folder_in_folder(
            title,
            shared_folder_id,
        )

        if len(stage_column_folder_names) > 0:
            # Create sub-folders
            for stage_column_folder in stage_column_folder_names:
                stage_column_folder_ids.append(
                    create_folder_in_folder(
                        stage_column_folder,
                        top_level_folder_id,
                    )
                )

        # Process rows in smaller batches
        batch_size = 10  # Adjust based on your needs
        max_workers = 5  # Reduced number of concurrent workers
        processed_rows = []

        for i in range(0, len(data), batch_size):
            batch = data[i:i + batch_size]
            with ThreadPoolExecutor(max_workers=max_workers) as executor:
                row_args = [(self, row, stage_column_index, stage_column_folder_ids, creds)
                    for row in batch]
                batch_results = list(executor.map(process_row, row_args))
                processed_rows.extend(batch_results)
                time.sleep(1)  # Add delay between batches

        # Add header and processed rows to columns
        columns = [keys] + processed_rows

        spreadsheet = {"properties": {"title": title}}
        spreadsheet = (
            service.spreadsheets()
            .create(body=spreadsheet, fields="spreadsheetId")
            .execute()
        )

        ss_id = spreadsheet.get("spreadsheetId")

        width_10 = chr(65 + len(columns[0]) % 26)
        width_1 = chr(64 + len(columns[0]) // 26) if len(columns[0]) > 25 else ''
        width = width_10 + width_1
        cell_range = f"Sheet1!A1:{width}{len(columns)}"
        logger.info(f"\n\nRange name: {cell_range} | {len(columns[0])} | {len(columns)}\n\n")
        body = {
                "values": columns
                }

        result = (
            service.spreadsheets()
            .values()
            .update(
                spreadsheetId=ss_id,
                range=cell_range,
                valueInputOption='USER_ENTERED',
                body=body,
            )
            .execute()
        )
        logger.info(f"{result.get('updatedCells')} cells created.")

        # Apply formatting
        service.spreadsheets().batchUpdate(
            spreadsheetId=ss_id,
            body={"requests": format_genesis_g_sheets(columns)}
        ).execute()

        # Move the document to shared folder
        file = {}
        if top_level_folder_id:
            file = (
                drive_service.files()
                .update(
                    fileId=ss_id,
                    addParents=top_level_folder_id,
                    fields="id, webViewLink, parents",
                )
                .execute()
            )
            logger.info(f"File moved to folder - File ID: {file['id']} | Folder ID {file['parents'][0]}")

        # Test only - read file contents to confirm write
        # results = read_g_sheet(ss_id, cell_range, creds)
        # logger.info(f"Results from storing, then reading sheet: {results}")

        folder_url = get_g_folder_web_link(top_level_folder_id, creds)
        file_url = file.get("webViewLink")

        g_file_version_data = {
            "g_file_id": ss_id,
            "g_file_name": title,
            "g_file_type": "sheet",
            "g_file_parent_id": top_level_folder_id,
            "g_file_size": None,
            "g_file_version": "1",
        }

        service._http.http.close()
        drive_service._http.http.close()

        insert_into_g_drive_file_version_table(self, g_file_version_data)

        return {"Success": True, "file_id": spreadsheet.get("spreadsheetId"), "file_url": file_url, "folder_url": folder_url}

    except HttpError as error:
        logger.info(f"An error occurred: {error}")
        return error

def create_g_sheet_v4(g_sheet_values, g_sheet_name = "Google Sheet", g_folder_id=None, creds=None) -> dict:
    """
    Create a Google Sheet with the given values.
    Load pre-authorized user credentials from the environment.
    """

    creds = load_creds()

    try:
        service = build("sheets", "v4", credentials=creds)

        # Create the Google Sheet
        spreadsheet = {"properties": {"title": g_sheet_name}}
        spreadsheet = (
            service.spreadsheets()
            .create(body=spreadsheet, fields="spreadsheetId")
            .execute()
        )

        ss_id = spreadsheet.get("spreadsheetId")
        logger.info(f"Spreadsheet ID: {ss_id}")

        # Prepare the body for the update request
        body = {
            "values": g_sheet_values
        }

        # Update the Google Sheet with the new values
        result = (
            service.spreadsheets()
            .values()
            .update(
                spreadsheetId=ss_id,
                range="Sheet1!A1",
                valueInputOption="USER_ENTERED",
                body=body,
            )
            .execute()
        )

        logger.info(f"{result.get('updatedCells')} cells created.")

        if folder_id:
            logger.info(f'Moving document to folder {folder_id}')
            service.files().update(fileId=ss_id, addParents=folder_id).execute()

        return {
            "Success": True,
            "file_id": spreadsheet.get("spreadsheetId"),
        }

    except Exception as e:
        logger.error(f"An error occurred: {str(e)}")
        return {
            "Success": False,
            "Error": str(e),
        }

def write_g_sheet_cell_v3(spreadsheet_id=None, cell_range=None, value=None, creds=None):
    logger.info(f"Entering write_g_sheet with ss_id: {spreadsheet_id}")

    creds = load_creds()
    service = build("drive", "v3", credentials=creds)
    result = read_g_sheet(spreadsheet_id, cell_range, creds)

    start_col, start_row, end_col, end_row, num_cells = (
        parse_cell_range(cell_range)
    )

    value_arr = value.split(",")

    if len(value_arr) != num_cells:
        raise ValueError("Number of values does not match the number of cells in cell_range")

    logger.info(f"Start Column: {start_col}, Start Row: {start_row}, End Column: {end_col}, End Row: {end_row}")

    # Update the result['cell_values'] with the values from value_arr
    index = 0
    for col in range(start_col - 1, end_col):
        for row in range(start_row - 1, end_row):
            result['cell_values'][row][col] = value_arr[index]
            index += 1

    # # Prepare the body for the update request
    # body = {
    #     "values": result['cell_values'][start_row - 1:end_row]
    # }

    # Update the Google Sheet with the new values - DOES NOT WORK ON SPCS DOCKER
    # service = build("sheets", "v4", credentials=creds)
    # result = (
    #     service.spreadsheets()
    #     .values()
    #     .update(
    #         spreadsheetId=spreadsheet_id,
    #         range=cell_range,
    #         valueInputOption='USER_ENTERED',
    #         body=body,
    #     )
    #     .execute()
    # )

    # Write the updated values back to the Google Sheet using openpyxl
    try:
        # Create a new workbook and worksheet
        new_workbook = openpyxl.Workbook()
        new_worksheet = new_workbook.active

        # Write the updated cell values to the new worksheet
        for row_idx, row in enumerate(result['cell_values'], start=1):
            for col_idx, cell_value in enumerate(row, start=1):
                new_worksheet.cell(row=row_idx, column=col_idx, value=cell_value)

        # Save the workbook to a temporary file
        temp_file_path = "temp_google_sheet.xlsx"
        new_workbook.save(temp_file_path)

        # Upload the file back to Google Drive
        # service = result['service'] #build("drive", "v3", credentials=creds)
        media = MediaFileUpload(temp_file_path, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        file = service.files().update(fileId=spreadsheet_id, media_body=media).execute()

        logger.info(f"File ID: {file.get('id')}")
        return {
            "Success": True,
            "updatedCells": result.get("updatedCells"),
            "file_id": file.get("id"),
        }

    except Exception as e:
        logger.info(f"An error occurred: {str(e)}")
        return {
            "Success": False,
            "Error": str(e),
        }


def write_g_sheet_cell_v4(
    spreadsheet_id=None, cell_range=None, value=None, creds=None
):
    creds = load_creds()

    service = build("sheets", "v4", credentials=creds)

    body = {"values": [[value]]}

    result = (
        service.spreadsheets()
        .values()
        .update(
            spreadsheetId=spreadsheet_id,
            range=cell_range,
            valueInputOption="USER_ENTERED",
            body=body,
        )
        .execute()
    )
    return {
        "Success": True,
        "updatedCells": result.get("updatedCells"),
    }


def read_g_sheet(spreadsheet_id=None, cell_range=None, creds=None) -> dict:
    """
    Reads the content of a Google Sheet.
    Load pre-authorized user credentials from the environment.
    """
    logger.info(f"Entering read_g_sheet with ss_id: {spreadsheet_id}")

    creds = load_creds()

    try:
        service = build("drive", "v3", credentials=creds)

        request = service.files().export_media(
            fileId=spreadsheet_id,
            mimeType="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        fh = BytesIO()
        downloader = MediaIoBaseDownload(fh, request)
        done = False
        while done is False:
            status, done = downloader.next_chunk()
            logger.info("Download %d%%" % int(status.progress() * 100))
        fh.seek(0)
        workbook = openpyxl.load_workbook(filename=fh, data_only=False)
        worksheet = workbook[workbook.sheetnames[0]]

        # Extract the content of the worksheet
        rows = []
        for row in worksheet.iter_rows(values_only=True):
            rows.append(list(row))

        # Parse and validate the cell range
        if cell_range:
            # Handle whole column ranges like A:A
            col_range_match = re.match(r"([A-Za-z]{1,2}):([A-Za-z]{1,2})", cell_range)
            if col_range_match:
                start_col, end_col = col_range_match.groups()
                start_col_num = column_to_number(start_col) - 1
                end_col_num = column_to_number(end_col) - 1

                # Get all rows but only the specified columns
                rows = [row[start_col_num:end_col_num + 1] for row in rows]

            else:
                # Handle normal ranges like A1:B2
                match = re.match(r"([A-Za-z]{1,2})(\d+):([A-Za-z]{1,2})(\d+)", cell_range)
                if match:
                    start_col, start_row, end_col, end_row = match.groups()

                    # Convert column letters to numbers (A=1, B=2, AA=27, etc)
                    start_col_num = column_to_number(start_col) - 1  # Convert to 0-based index
                    end_col_num = column_to_number(end_col) - 1

                    # Convert row numbers to 0-based indices
                    start_row_num = int(start_row) - 1
                    end_row_num = int(end_row) - 1

                    # Ensure valid ranges
                    start_col_num = max(0, start_col_num)
                    end_col_num = min(len(rows[0])-1 if rows else 0, end_col_num)
                    start_row_num = max(0, start_row_num)
                    end_row_num = min(len(rows)-1 if rows else 0, end_row_num)

                    # Filter rows based on range
                    rows = rows[start_row_num:end_row_num + 1]

                    # Filter columns for each row
                    rows = [row[start_col_num:end_col_num + 1] for row in rows]
                else:
                    # Handle single cell case (e.g. "A1" or "AA1")
                    match = re.match(r"([A-Za-z]{1,2})(\d+)", cell_range)
                    if match:
                        col, row = match.groups()
                        col_num = column_to_number(col) - 1
                        row_num = int(row) - 1

                        # Ensure valid indices
                        if 0 <= row_num < len(rows) and 0 <= col_num < len(rows[0]):
                            rows = [[rows[row_num][col_num]]]
                        else:
                            rows = [[None]]

        service._http.http.close()

        return {
            "Success": True,
            "cell_values": rows,
        }
    except Exception as error:
        logger.error(f"HTTPError in read sheet: {error} - {spreadsheet_id}")
        return {"Success": False,"error": str(error)}


def delete_g_file(file_id=None, creds=None) -> dict:
    """
    Deletes a Google Sheet.
    Load pre-authorized user credentials from the environment.
    """
    logger.info(f"Entering delete_g_file with file_id: {file_id}")

    creds = load_creds()

    try:
        service = build("drive", "v3", credentials=creds)

        service.files().delete(fileId=file_id).execute()

        service._http.http.close()

        return {"Success": True, "message": f"File {file_id} deleted successfully"}
    except Exception as error:
        logger.info(f"HTTPError in read sheet: {error} - {file_id}")
        return {"Success": False,"error": error}
