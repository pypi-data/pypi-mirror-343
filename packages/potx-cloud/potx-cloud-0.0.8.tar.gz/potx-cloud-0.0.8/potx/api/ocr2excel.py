# -*- coding: UTF-8 -*-
"""
开通和使用👉[免费教程](https://www.bilibili.com/video/BV13J4m1s7L7/)
"""
import json
from collections import OrderedDict
from pathlib import Path

import pandas as pd
from loguru import logger
from pofile import get_files, mkdir
from poprogress import simple_progress
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Alignment

import potx
from potx.api.ocr import *


def VatInvoiceOCR2Excel(input_path, output_path=r'./', output_excel='VatInvoiceOCR2Excel.xlsx', img_url=None,
                        configPath=None, id=None, key=None, file_name=False, trans=False):
    """
    将OCR识别的增值税发票数据转换为Excel表格。

    该函数主要处理从图像文件中提取的增值税发票数据，通过OCR技术识别后，将数据整理并输出到Excel表格中。
    这对于财务人员自动整理和核对发票信息非常有用。

    :param input_path: 输入文件路径，可以是单个文件或文件夹
    :param output_path: 输出Excel文件的路径，默认为None，表示使用函数默认文件名并保存在当前目录
    :param output_excel: 输出Excel文件的名称，默认为'VatInvoiceOCR2Excel.xlsx'
    :param img_url: 图像文件的URL地址，用于远程处理
    :param configPath: 配置文件路径，用于指定OCR引擎的配置
    :param id: OCR引擎的用户ID
    :param key: OCR引擎的用户密钥
    :param file_name: 是否在Excel中包含文件名作为一行数据，默认为False
    :param trans: 是否进行数据转换，默认为False。如果设置为True，将尝试将识别到的文本数据转换为相应的数字或日期格式
    """

    vat_img_files = get_files(input_path)
    if vat_img_files is None:
        raise BaseException(f'{input_path}这个文件目录下，没有存放任何发票，请确认后重新运行')
    mkdir(Path(output_path).absolute())  # 如果不存在，则创建输出目录
    if output_excel.endswith('.xlsx') or output_excel.endswith('xls'):  # 如果指定的输出excel结尾不正确，则报错退出
        abs_output_excel = Path(output_path).absolute() / output_excel
    else:  # 指定了，但不是xlsx或者xls结束
        raise BaseException(
            f'输出结果名：output_excel参数，必须以xls或者xlsx结尾，您的输入:{output_excel}有误，请修改后重新运行')
    res_df = []  # 装全部识别的结果
    for vat_img in simple_progress(vat_img_files):
        try:
            api_res = VatInvoiceOCR(file_path=str(vat_img), file_url=img_url, configPath=configPath, id=id, key=key)
            # 处理api_res可能是列表的情况（多页pdf）
            if isinstance(api_res, list):
                # 遍历每一页结果
                for page_res in api_res:
                    VatInvoiceOCR2(page_res, vat_img, file_name, trans, res_df)
            else:
                VatInvoiceOCR2(api_res, vat_img, file_name, trans, res_df)
        except Exception as e:
            logger.error(e)
    # 整理全部识别结果
    if len(res_df) > 0:
        # 将所有结果转换位DataFrame并保存
        df = pd.DataFrame(res_df)
        df.to_excel(str(abs_output_excel), index=None, engine='openpyxl')
    else:
        logger.warning(f'该文件夹下，没有任何符合条件的发票图片/PDF文件')


def BankSlipOCR2Excel(input_path, output_path=r'./', output_excel='BankSlipOCR2Excel.xlsx', img_url=None,
                      configPath=None, id=None, key=None):
    """
    将OCR识别的增值税发票数据转换为Excel表格。

    该函数主要处理从图像文件中提取的增值税发票数据，通过OCR技术识别后，将数据整理并输出到Excel表格中。
    这对于财务人员自动整理和核对发票信息非常有用。

    :param input_path: 输入文件路径，可以是单个文件或文件夹
    :param output_path: 输出Excel文件的路径，默认为None，表示使用函数默认文件名并保存在当前目录
    :param output_excel: 输出Excel文件的名称，默认为'BankSlipOCR2Excel.xlsx'
    :param img_url: 图像文件的URL地址，用于远程处理
    :param configPath: 配置文件路径，用于指定OCR引擎的配置
    :param id: OCR引擎的用户ID
    :param key: OCR引擎的用户密钥
    :param file_name: 是否在Excel中包含文件名作为一行数据，默认为False
    :param trans: 是否进行数据转换，默认为False。如果设置为True，将尝试将识别到的文本数据转换为相应的数字或日期格式
    """

    vat_img_files = get_files(input_path)
    if vat_img_files is None:
        raise BaseException(f'{input_path}这个文件目录下，没有存放任何发票，请确认后重新运行')
    mkdir(Path(output_path).absolute())  # 如果不存在，则创建输出目录
    if output_excel.endswith('.xlsx') or output_excel.endswith('xls'):  # 如果指定的输出excel结尾不正确，则报错退出
        abs_output_excel = Path(output_path).absolute() / output_excel
    else:  # 指定了，但不是xlsx或者xls结束
        raise BaseException(
            f'输出结果名：output_excel参数，必须以xls或者xlsx结尾，您的输入:{output_excel}有误，请修改后重新运行')
    res_df = []  # 装全部识别的结果
    for vat_img in simple_progress(vat_img_files):
        try:
            api_res = BankSlipOCR(file_path=str(vat_img), file_url=img_url, configPath=configPath, id=id, key=key)
            api_res_json = json.loads(str(api_res), object_pairs_hook=OrderedDict)
            bank_infos = api_res_json['BankSlipInfos']
            dict_pandas = {}  # 存放一行数据
            info_i = 0
            for info in bank_infos:
                info_i = info_i + 1
                name = info['Name']
                value = info['Value']
                if name == '机构' and info_i != 1:
                    d_p = pd.DataFrame(dict_pandas, index=[0])
                    res_df.append(d_p)
                    dict_pandas = {}  # 存放一行数据

                dict_pandas[name] = value
            if dict_pandas:
                d_p = pd.DataFrame(dict_pandas, index=[0])
                res_df.append(d_p)

        except Exception as e:
            logger.error(e)
    # 整理全部识别结果
    if len(res_df) > 0:
        res_excel = res_df[0]
        for index, line_df in enumerate(res_df):
            if index == 0:
                continue
            res_excel = res_excel._append(line_df)
        pd.DataFrame(res_excel).to_excel(str(abs_output_excel))  # 写入Excel
    else:
        logger.warning(f'该文件夹下，没有任何符合条件的发票图片/PDF文件')


def IDCardOCR2Excel(input_path=None, output_path=None, output_excel='IDCardOCR2Excel.xlsx', img_url=None,
                    configPath=None, id=None, key=None):
    """
    批量识别身份证，并保存在Excel中
    :param input_path: 身份证存放位置，可以填单个文件，也可以填一个目录
    :param output_path:
    :param output_excel:
    :param img_url:
    :param configPath:
    :return:
    """
    if input_path is None and img_url is None:
        raise BaseException(f'参数异常,请检查后重新运行!')
    id_img_files = [img_url] if input_path is None else get_files(input_path)
    if id_img_files is None or len(id_img_files) == 0:
        raise BaseException(f'{input_path}这个文件目录下，没有存放任何身份证，请确认后重新运行')
    output_path = output_path or './'
    mkdir(Path(output_path).absolute())  # 如果不存在，则创建输出目录
    if output_excel.endswith('.xlsx') or output_excel.endswith('xls'):  # 如果指定的输出excel结尾不正确，则报错退出
        abs_output_excel = Path(output_path).absolute() / output_excel
    else:  # 指定了，但不是xlsx或者xls结束
        raise BaseException(
            f'输出结果名：output_excel参数，必须以xls或者xlsx结尾，您的输入:{output_excel}有误，请修改后重新运行')
    res_df = []  # 装全部识别的结果
    for id_img in simple_progress(id_img_files):
        try:
            api_res = IDCardOCR(file_path=str(id_img), file_url=img_url, configPath=configPath, id=id, key=key)
            api_res_json = json.loads(str(api_res))
            del api_res_json['ReflectDetailInfos']

            res_df.append(pd.DataFrame(api_res_json, index=[0]))
        except Exception as e:
            logger.error(e)
            continue
    # 整理全部识别结果
    if len(res_df) > 0:
        res_excel = res_df[0]
        for index, line_df in enumerate(res_df):
            if index == 0:
                continue
            res_excel = res_excel._append(line_df)
        pd.DataFrame(res_excel).to_excel(str(abs_output_excel))  # 写入Excel
    else:
        logger.error(f'该文件夹下，没有任何符合条件的身份证图片')


def TrainTicketOCR2Excel(input_path: str = None, output_path=None, output_excel: str = 'TrainTicketOCR2Excel.xlsx',
                         img_url: str = None, configPath: str = None, id=None, key=None) -> None:
    if input_path is None and img_url is None:
        raise BaseException(f'参数异常,请检查后重新运行!')
    ticket_files = [img_url] if input_path is None else get_files(input_path)
    if ticket_files is None or len(ticket_files) == 0:
        raise BaseException(f'{input_path}这个文件目录下，没有存放任何火车票，请确认后重新运行')
    mkdir(Path(output_path).absolute())
    # 检查输出文件名是否以.xlsx或.xls结尾，如果不是，抛出异常
    if output_excel.endswith('.xlsx') or output_excel.endswith('xls'):  # 如果指定的输出excel结尾不正确，则报错退出
        abs_output_excel = Path(output_path).absolute() / output_excel
    else:  # 指定了，但不是xlsx或者xls结束
        raise BaseException(
            f'输出结果名：output_excel参数，必须以xls或者xlsx结尾，您的输入:{output_excel}有误，请修改后重新运行')
    ticket_list = []
    for ticket in simple_progress(ticket_files):
        try:
            api_res = TrainTicketOCR(file_path=ticket, file_url=img_url, configPath=configPath, id=id, key=key)
            api_res_json = json.loads(str(api_res))
            ticket_list.append(pd.DataFrame(api_res_json, index=[0]))
        except Exception:
            continue
    # 整理全部识别结果
    if len(ticket_list) > 0:
        res_excel = ticket_list[0]
        for index, line_df in enumerate(ticket_list):
            if index == 0:
                continue
            res_excel = res_excel._append(line_df)
        pd.DataFrame(res_excel).to_excel(str(str(abs_output_excel)))  # 写入Excel
    else:
        logger.error(f'该文件夹下，没有任何符合条件的火车票图片')


def BankCardOCR2Excel(input_path, output_path=None, output_excel='BankCardOCR2Excel.xlsx', img_url=None,
                      configPath=None, id=None, key=None):
    """
    识别银行卡，自动保存为Excel文件
    :param input_path: 必填，银行卡图片的位置
    :param output_path: 选填，输出Excel的位置
    :param output_excel: 选填，输出Excel的名称
    :param img_url: 选填，可以是网络图片
    :param configPath: 已废弃
    :param id: 你的腾讯账号的密钥，获取方式：https://curl.qcloud.com/fuOGcm2R
    :param key: 你的腾讯账号的密钥，获取方式：https://curl.qcloud.com/fuOGcm2R
    :return:
    """
    test_json = potx.ocr.BankCardOCR(file_path=input_path, file_url=img_url, configPath=configPath, id=id,
                                     key=key)
    df = pd.DataFrame(json.loads(str(test_json)), index=[0])
    if output_path is None:
        output_path = './'
    mkdir(Path(output_path).absolute())  # 如果不存在，则创建输出目录
    if output_excel.endswith('.xlsx') or output_excel.endswith('xls'):  # 如果指定的输出excel结尾不正确，则报错退出
        abs_output_excel = Path(output_path).absolute() / output_excel
    df.to_excel(str(abs_output_excel), index=False)


def LicensePlateOCR2Excel(input_path=None, output_path=None, output_excel='LicensePlateOCR2Excel.xlsx', img_url=None,
                          configPath=None, id=None, key=None):
    """
    识别银行卡，自动保存为Excel文件
    :param input_path: 必填，银行卡图片的位置
    :param output_path: 选填，输出Excel的位置
    :param output_excel: 选填，输出Excel的名称
    :param img_url: 选填，可以是网络图片
    :param configPath: 已废弃
    :param id: 你的腾讯账号的密钥，获取方式：https://curl.qcloud.com/fuOGcm2R
    :param key: 你的腾讯账号的密钥，获取方式：https://curl.qcloud.com/fuOGcm2R
    :return:
    """
    if input_path is None and img_url is None:
        raise BaseException(f'参数异常,请检查后重新运行!')
    img_paths = [img_url] if input_path is None else get_files(input_path)
    if img_paths is None or len(img_paths) == 0:
        raise BaseException(f'{input_path}目录下,无银行卡,请检查后重新运行.')
    output_path = output_path or './'
    mkdir(Path(output_path).absolute())  # 如果不存在，则创建输出目录
    if output_excel.endswith('.xlsx') or output_excel.endswith('xls'):  # 如果指定的输出excel结尾不正确，则报错退出
        abs_output_excel = Path(output_path).absolute() / output_excel
    else:
        raise BaseException(
            f'输出结果名：output_excel参数，必须以xls或者xlsx结尾，您的输入:{output_excel}有误，请修改后重新运行')
    res_df = []
    for item in simple_progress(img_paths):
        try:
            api_res = potx.ocr.LicensePlateOCR(file_path=input_path, file_url=img_url, configPath=configPath, id=id,
                                               key=key)
            api_res_json = json.loads(str(api_res))
            res_df.append(api_res_json)
        except Exception as e:
            logger.error(f'{item}识别失败，原因：{e}')
    df = pd.DataFrame(res_df)
    df.to_excel(str(abs_output_excel), index=False)


def BizLicenseOCR2Excel(input_path=None, output_path=r'./', output_excel='BizLicenseOCR2Excel.xlsx', img_url=None,
                        configPath=None, id=None, key=None):
    """
    将营业执照OCR识别结果整理并保存到Excel中。

    :param input_path: 图片输入路径，包含营业执照图片。
    :param output_path: 输出路径，默认为当前目录。
    :param output_excel: 输出的Excel文件名，默认为'BizLicenseOCR2Excel.xlsx'。
    :param img_url: 图片URL，如果提供，将通过URL进行识别。
    :param configPath: 配置文件路径，用于指定OCR识别的配置。
    :param id: API的用户ID。
    :param key: API的密钥。
    :param file_name: 是否在结果中包含文件名，默认为False。
    :param trans: 是否进行翻译，默认为False。
    :raises BaseException: 当输入目录为空或输出文件名格式不正确时抛出异常。
    """
    if input_path is None and img_url is None:
        raise BaseException(f'参数异常,请检查后重新运行!')
    vat_img_files = [img_url] if input_path is None else get_files(input_path)
    # 如果文件列表为空，抛出异常
    if vat_img_files is None or len(vat_img_files) == 0:
        raise BaseException(f'{input_path}这个文件目录下，没有存放任何营业执照，请确认后重新运行')
    # 创建输出路径
    mkdir(Path(output_path).absolute())  # 如果不存在，则创建输出目录
    # 检查输出文件名是否以.xlsx或.xls结尾，如果不是，抛出异常
    if output_excel.endswith('.xlsx') or output_excel.endswith('xls'):  # 如果指定的输出excel结尾不正确，则报错退出
        abs_output_excel = Path(output_path).absolute() / output_excel
    else:  # 指定了，但不是xlsx或者xls结束
        raise BaseException(
            f'输出结果名：output_excel参数，必须以xls或者xlsx结尾，您的输入:{output_excel}有误，请修改后重新运行')
    # 初始化结果数据框
    res_df = []  # 装全部识别的结果
    # 对每个营业执照图片进行识别
    for vat_img in simple_progress(vat_img_files):
        try:
            # 调用OCR API进行识别
            api_res = BizLicenseOCR(file_path=str(vat_img), file_url=img_url, configPath=configPath, id=id, key=key)
            # 将识别结果转换为JSON格式并添加到结果数据框
            api_res_json = json.loads(str(api_res))
            res_df.append(api_res_json)
        except Exception as e:
            # 打印识别失败的信息
            logger.error(f'{vat_img}识别失败，原因：{e}')
    # 将所有识别结果合并成一个数据框
    biz_def = pd.DataFrame(res_df)
    # 将结果数据框保存到Excel文件
    biz_def.to_excel(str(abs_output_excel), index=None)


def RecognizeGeneralInvoiceOCR2Excel(input_path=None, output_path=None,
                                     output_excel='RecognizeGeneralInvoiceOCR2Excel.xlsx',
                                     img_url=None,
                                     configPath=None, id=None, key=None, sub_type=None):
    """
   通用识别发票信息并保存到Excel中。

   :param input_path: 图片输入路径。
   :param output_path: 输出路径，默认为当前目录。
   :param output_excel: 输出的Excel文件名，默认为'BizLicenseOCR2Excel.xlsx'。
   :param img_url: 图片URL，如果提供，将通过URL进行识别。
   :param configPath: 配置文件路径，用于指定OCR识别的配置。
   :param id: API的用户ID。
   :param key: API的密钥。
   :param sub_type: 识别的类型 具体参数详见:https://cloud.tencent.com/document/api/866/90802  中SubType的列表信息
   :raises BaseException: 当输入目录为空或输出文件名格式不正确时抛出异常。
   """
    if input_path is None and img_url is None:
        raise BaseException(f'参数异常,请检查后重新运行!')
    img_paths = [img_url] if input_path is None else get_files(input_path)
    if img_paths is None or len(img_paths) == 0:
        raise BaseException(f'{input_path}目录下,无机动车发票,请检查后重新运行.')
    output_path = output_path or './'
    mkdir(Path(output_path).absolute())  # 如果不存在，则创建输出目录
    if output_excel.endswith('.xlsx') or output_excel.endswith('xls'):  # 如果指定的输出excel结尾不正确，则报错退出
        abs_output_excel = Path(output_path).absolute() / output_excel
    else:  # 指定了，但不是xlsx或者xls结束
        raise BaseException(
            f'输出结果名：output_excel参数，必须以xls或者xlsx结尾，您的输入:{output_excel}有误，请修改后重新运行')
    try:
        for item in simple_progress(img_paths):
            api_res = RecognizeGeneralInvoice(file_path=str(item), file_url=img_url, configPath=configPath, id=id,
                                              key=key)
            api_res_json = json.loads(str(api_res))
            invoice_items = api_res_json['MixedInvoiceItems']
            # 创建 Excel 工作簿
            wb = Workbook()
            if 'Sheet' in wb.sheetnames:
                del wb['Sheet']  # 删除默认创建的Sheet
            for invoice_item in invoice_items:
                rows = []
                if invoice_item['Code'] != 'OK':
                    continue
                if sub_type is not None and invoice_item['Type'] != sub_type:
                    continue
                single_invoice_info = invoice_item['SingleInvoiceInfos']
                single_invoice_infos = single_invoice_info[invoice_item['SubType']]
                list_fields = {k: v for k, v in single_invoice_infos.items() if
                               isinstance(single_invoice_infos[k], list)}
                non_list_fields = {k: v for k, v in single_invoice_infos.items() if
                                   not isinstance(single_invoice_infos[k], list)}
                if len(list_fields) == 0:
                    rows.append({**non_list_fields})
                else:
                    for list_name, records in list_fields.items():
                        if len(records) == 0:
                            row = {**non_list_fields}
                            rows.append(row)
                            continue
                        for record in records:
                            row = {**non_list_fields, **record}
                            rows.append(row)

                # 转为 DataFrame
                df = pd.DataFrame(rows)

                # 获取主字段（不包含列表字段）
                main_fields = list(non_list_fields.keys())
                # 获取所有字段（含列表字段 + 主字段）
                all_fields = list(df.columns)

                sheet_name = invoice_item['TypeDescription']
                if sheet_name not in wb.sheetnames:
                    ws = wb.create_sheet(sheet_name)
                else:
                    ws = wb[sheet_name]
                ws.title = invoice_item['TypeDescription']

                ws.append(all_fields)

                for i, row in enumerate(dataframe_to_rows(df, index=False, header=False), start=2):
                    ws.append(row)

                # 合并主字段的单元格（按每组记录数来合并）
                record_counts = {}
                for list_name, records in list_fields.items():
                    record_counts[list_name] = len(records)

                # 每次合并这些字段在主表区域
                merge_start_row = 2
                for list_name, count in record_counts.items():
                    merge_end_row = merge_start_row + count - 1
                    for col_name in main_fields:
                        col_idx = all_fields.index(col_name) + 1  # Excel列是1-based
                        if count > 1:
                            ws.merge_cells(start_row=merge_start_row, start_column=col_idx,
                                           end_row=merge_end_row, end_column=col_idx)
                            # 垂直居中
                            cell = ws.cell(row=merge_start_row, column=col_idx)
                            cell.alignment = Alignment(vertical="center")
                    merge_start_row = merge_end_row + 1

                # 保存
            wb.save((abs_output_excel))
    except Exception as e:
        logger.error(f'{item}识别失败，原因：{e}')


def VatInvoiceOCR2(api_res, vat_img, file_name, trans, res_df):
    """
    处理单个结果
    Args:
        api_res: 识别结果
        vat_img: 图片路径
        file_name: 是否包含文件铭
        trans: 是否进行数据转换
        res_df: 结果列表，用于村粗处理后的数据

    Returns:

    """
    try:
        api_res_json = json.loads(str(api_res))
        # 处理VatInvoiceInfos部分
        if 'VatInvoiceInfos' in api_res_json:
            VatInvoiceInfos = api_res_json['VatInvoiceInfos']
            dict_pandas = {}

            # add文件铭
            if file_name:
                dict_pandas['文件名'] = Path(vat_img).name

            # 处理备注字段
            beizhu_value = ''

            # 处理所有发票
            for item in VatInvoiceInfos:
                if item['Name'] == '备注':
                    beizhu_value += item['Value']
                else:
                    if trans:
                        pass
                    else:
                        dict_pandas[item['Name']] = item['Value']
            # 添加备注字段
            dict_pandas['备注'] = beizhu_value

            # 处理Items部分
            if 'Items' in api_res_json:
                field_mapping = {
                    'Name': '商品名称',
                    'Spec': '规格型号',
                    'Unit': '单位',
                    'Quantity': '数量',
                    'UnitPrice': '单价',
                    'AmountWithoutTax': '金额'}

                Items = api_res_json['Items']
                if Items and len(Items) > 0:
                    dict_pandas['有商品明细'] = '是'

                    # 将第一个商品信息添加到结果
                    first_item = Items[0]
                    for key, value in first_item.items():
                        column_name = f"{field_mapping.get(key, key)}"
                        dict_pandas[column_name] = value
                else:
                    dict_pandas['有商品明细'] = '否'
            # 将处理好的结果添加到结果列表
            res_df.append(dict_pandas)

            logger.info(f"成功处理文件: {vat_img}")

    except Exception as e:
        logger.error(f"处理文件 {vat_img} 时出错: {str(e)}")


def DriverLicenseOCR2Excel(input_path=None, output_path=None, output_excel='DriverLicenseOCR2Excel.xlsx', img_url=None,
                           configPath=None, id=None, key=None):
    """
    将驾驶证OCR识别结果整理并保存到Excel中。

    :param input_path: 图片输入路径，包含驾驶证图片。
    :param output_path: 输出路径，默认为当前目录。
    :param output_excel: 输出的Excel文件名，默认为'DriverLicenseOCR2Excel.xlsx'。
    :param img_url: 图片URL，如果提供，将通过URL进行识别。
    :param configPath: 配置文件路径，用于指定OCR识别的配置。
    :param id: API的用户ID。
    :param key: API的密钥。
    :raises BaseException: 当输入目录为空或输出文件名格式不正确时抛出异常。
    """
    if input_path is None and img_url is None:
        raise BaseException(f'参数异常,请检查后重新运行!')
    vat_img_files = [img_url] if input_path is None else get_files(input_path)
    # 如果文件列表为空，抛出异常
    if vat_img_files is None or len(vat_img_files) == 0:
        raise BaseException(f'{input_path}这个文件目录下，没有存放任何营业执照，请确认后重新运行')
    output_path = output_path or './'
    # 创建输出路径
    mkdir(Path(output_path).absolute())  # 如果不存在，则创建输出目录
    # 检查输出文件名是否以.xlsx或.xls结尾，如果不是，抛出异常
    if output_excel.endswith('.xlsx') or output_excel.endswith('xls'):  # 如果指定的输出excel结尾不正确，则报错退出
        abs_output_excel = Path(output_path).absolute() / output_excel
    else:  # 指定了，但不是xlsx或者xls结束
        raise BaseException(
            f'输出结果名：output_excel参数，必须以xls或者xlsx结尾，您的输入:{output_excel}有误，请修改后重新运行')
    # 初始化结果数据框
    res_df = []  # 装全部识别的结果
    # 对每个营业执照图片进行识别
    for vat_img in simple_progress(vat_img_files):
        try:
            # 调用OCR API进行识别
            api_res = DriverLicenseOCR(file_path=str(vat_img), file_url=img_url, configPath=configPath, id=id, key=key)
            # 将识别结果转换为JSON格式并添加到结果数据框
            api_res_json = json.loads(str(api_res))
            if 'RecognizeWarnCode' in api_res_json:
                del api_res_json['RecognizeWarnCode']
            if 'RecognizeWarnMsg' in api_res_json:
                del api_res_json['RecognizeWarnMsg']
            res_df.append(api_res_json)
        except Exception as e:
            # 打印识别失败的信息
            logger.error(f'{vat_img}识别失败，原因：{e}')
    # 将所有识别结果合并成一个数据框
    biz_def = pd.DataFrame(res_df)
    # 将结果数据框保存到Excel文件
    biz_def.to_excel(str(abs_output_excel), index=None)


def VehicleLicenseOCR2Excel(input_path=None, output_path=None, output_excel='VehicleLicenseOCR2Excel.xlsx',
                            img_url=None,
                            configPath=None, id=None, key=None):
    """
    将行驶证OCR识别结果整理并保存到Excel中。

    :param input_path: 图片输入路径，包含行驶证图片。
    :param output_path: 输出路径，默认为当前目录。
    :param output_excel: 输出的Excel文件名，默认为'VehicleLicenseOCR2Excel.xlsx'。
    :param img_url: 图片URL，如果提供，将通过URL进行识别。
    :param configPath: 配置文件路径，用于指定OCR识别的配置。
    :param id: API的用户ID。
    :param key: API的密钥。
    :raises BaseException: 当输入目录为空或输出文件名格式不正确时抛出异常。
    """
    if input_path is None and img_url is None:
        raise BaseException(f'参数异常,请检查后重新运行!')
    vat_img_files = [img_url] if input_path is None else get_files(input_path)
    # 如果文件列表为空，抛出异常
    if vat_img_files is None or len(vat_img_files) == 0:
        raise BaseException(f'{input_path}这个文件目录下，没有存放任何营业执照，请确认后重新运行')
    output_path = output_path or './'
    # 创建输出路径
    mkdir(Path(output_path).absolute())  # 如果不存在，则创建输出目录
    # 检查输出文件名是否以.xlsx或.xls结尾，如果不是，抛出异常
    if output_excel.endswith('.xlsx') or output_excel.endswith('xls'):  # 如果指定的输出excel结尾不正确，则报错退出
        abs_output_excel = Path(output_path).absolute() / output_excel
    else:  # 指定了，但不是xlsx或者xls结束
        raise BaseException(
            f'输出结果名：output_excel参数，必须以xls或者xlsx结尾，您的输入:{output_excel}有误，请修改后重新运行')
    # 初始化结果数据框
    res_df = []  # 装全部识别的结果
    # 对每个营业执照图片进行识别
    for vat_img in simple_progress(vat_img_files):
        try:
            # 调用OCR API进行识别
            api_res = VehicleLicenseOCR(file_path=str(vat_img), file_url=img_url, configPath=configPath, id=id, key=key)
            # 将识别结果转换为JSON格式并添加到结果数据框
            api_res_json = json.loads(str(api_res))
            front_info = api_res_json.get('FrontInfo')
            if front_info is None:
                raise BaseException('识别有异常')
            front_info['识别结果'] = "行驶证正面"
            res_df.append(front_info)
            back_info = api_res_json.get('BackInfo')
            if back_info is not None:
                back_info['识别结果'] = "行驶证背面"
                res_df.append(back_info)
        except Exception as e:
            # 打印识别失败的信息
            logger.error(f'{vat_img}识别失败，原因：{e}')
    # 将所有识别结果合并成一个数据框
    biz_def = pd.DataFrame(res_df)
    # 将结果数据框保存到Excel文件
    biz_def.to_excel(str(abs_output_excel), index=None)
