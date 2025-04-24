#!/usr/bin/env python
# -*- coding: utf-8 -*-
import re
from openpyxl import load_workbook
from collections import OrderedDict
from os import listdir
from .opc_obj import Generic, approve_name_and_register_guid
from . import opc_vars


def make_lib_file(workbook, lib_name:str, lib_dict:dict):
	output = "#!/usr/bin/env python\n# -*- encoding: utf-8 -*-\n"
	output += "from . import opc_obj"
	for lib in list(set(lib_dict.values())):
		output += ", " + lib
	output += """
	
class Gen_OPC_Obj(opc_obj.Generic):
	def test(self):
		print('Is ' + self.__class__.__name__)
		
	def _transform(self, diag=False):
		if diag: print("Already transformed into " + str(self.__class__))
		return self
	"""
	for sheet_name in workbook.sheetnames:
		output += "\nclass " + sheet_name + "(Gen_OPC_Obj):\n\n" + "\tdef __init__(self,opc_path, predecessor=None, description=u'', sig_range=None):"
		lib_dict[sheet_name] = lib_name
		ws = workbook[sheet_name]
		opc_children = []
		output += "\n\t\tself.opc_path = opc_path"
		output += "\n\t\tself.description = description"
		output += "\n\t\tself.sig_range = sig_range"
		output += "\n\t\tif predecessor is None:"
		for row in range(2,ws.max_row+1):
			if ws.cell(row=row, column=1).value is None:
				break
			sig_range = {}
			description = None
			parameter = None
			attribute_name = ws.cell(row=row, column=1).value
			if attribute_name.title() == 'Dummy':
				continue
			child_class_name = ws.cell(row=row, column=2).value
			col_6 = ws.cell(row=row, column=6).value
			#col_7 = ws.cell(row=row, column=7).value
			#col_8 = ws.cell(row=row, column=8).value
			#col_9 = ws.cell(row=row, column=9).value
			#col_10 = ws.cell(row=row, column=10).value
			#col_11 = ws.cell(row=row, column=11).value
			if isinstance(col_6, str): description = col_6 + u' '
			#if isinstance(col_7,str): description = col_7 + u' '
			#if isinstance(col_8,str): parameter = col_8
			#if isinstance(col_9,str): sig_range['Unit'] = col_9
			#if not col_10 is None: sig_range['Min'] = col_10
			#if not col_11 is None: sig_range['Max'] = col_11

			if 'string[' in child_class_name.lower():
				child_class_name = 'String'
			if child_class_name.title() in ['Real','Bool','Time','Uint','Dint','Dword','Word','Date_And_Time','String']:
				child_class_name = child_class_name.title()
			try:
				child_lib_name = lib_dict[child_class_name]
			except KeyError:
				print("Unknown type: " + child_class_name)
				print("At creation of: " + sheet_name)
				exit()

			opc_children.append(attribute_name)
			if child_lib_name == lib_name:
				output += "\n\t\t\tself." + attribute_name + " = " + child_class_name
			else:
				output += "\n\t\t\tself." + attribute_name + " = " + child_lib_name + "." + child_class_name
			if not description is None:
				output += "(opc_path + '." +attribute_name+ "', description=description + u'" + description + "'"
			else:
				output += "(opc_path + '." +attribute_name+ "', description=description"

			# if not parameter is None:
			# 	if parameter == '[para]':
			# 		output += ", parameter= self.parameter"
			# 	else:
			# 		output += ", parameter=u'" + parameter + "'"
			# else:
			# 	output += ", parameter=parameter"
			# if child_class_name == 'Bool':
			# 	pass
			# elif not sig_range == {}:
			# 	output += ", sig_range=" + str(sig_range)
			# else:
			# 	output += ", sig_range=sig_range"

			output += ")"

		output += "\n\t\t\tself.opc_children = " + str(opc_children)
		output += """		
		else:
			for attribute in [a for a in dir(predecessor) if not a.startswith('__') and not callable(getattr(predecessor,a))]:
				setattr(self,attribute,getattr(predecessor,attribute))
				"""
	return output, lib_dict

def update_init_file():
	with open('opc_class_lib/__init__.py', 'w', encoding='utf-8') as initFile:
		new_string = "__all__ = ["
		for file in listdir('opc_class_lib'):
			if "__init__.py" == file: continue
			if file[-3:] == '.py':
				new_string += "'" + file[:-3] + "',"
		new_string = new_string[:-1] + "]"
		initFile.write(new_string)

def create_from_StartValuesData():
	"""Function to read data from ABB 800M
	Start Value Analyzer files. The init-value
	is put in a .init_value"""
	dir_path = listdir('../../Input')
	start_value_folder_paths = ['Input\\' + folder_name for folder_name in dir_path if 'StartValuesData' in folder_name]
	if len(start_value_folder_paths) == 0:
		raise Exception("No folder name containing 'StartValuesData' in 'Input' folder")
	if len(start_value_folder_paths) == 1:
		print("Only one folder available '" + start_value_folder_paths[0] + "'.")
		print("Selected '" + start_value_folder_paths[0] + "' to read from.")
		file_paths = [start_value_folder_paths[0] + '\\' + file_path for file_path in listdir(start_value_folder_paths[0])]
	else:
		for idx, name in enumerate(start_value_folder_paths):
			print(str(idx + 1).ljust(5) + name)
		idx = -1
		while not 0 <= int(idx) < len(start_value_folder_paths):
			idx = input("Chose a folder-server to read from [1-" + str(len(start_value_folder_paths)) + "] input 'a' to abort:")
			if idx == r'a':
				raise Exception('User aborted')
			idx = int(idx)-1
		print("Selected '" + start_value_folder_paths[idx] + "' to read from.")
		file_paths = [start_value_folder_paths[idx] + '\\' + file_path for file_path in listdir(start_value_folder_paths[idx])]
	pattern = re.compile(r"^(.*?)\t(.*?)\t(.*?)\t(.*?)\t(.*?)$")
	root = Generic(None)
	for file_path in file_paths:
		if 'SessionLog' in file_path: continue
		with open(file_path,'r') as file:
			next(file) #Skipping the first row
			for line in file:
				match = pattern.match(line)
				if match:
					path, init_value, live_value, quality, value_type = match.groups()
					pos = root
					part_path = ''
					for part in path.split('.')[:-1]:
						part_path += '.' + part
						attr_name = approve_name_and_register_guid(pos, None, item_name=part)
						if hasattr(pos, attr_name):
							pos = getattr(pos, attr_name)
						else:
							new_child = Generic(part_path[1:])
							pos.opc_children.append(attr_name)
							setattr(pos, attr_name, Generic(part_path[1:]))
							pos = new_child
					attr_name = approve_name_and_register_guid(pos,None,item_name=path.split('.')[-1])
					if hasattr(pos,attr_name):
						leaf = getattr(pos,attr_name)
					else:
						leaf = opc_vars.OpcVariable(path)
						leaf.name_prop = {'Item Type Name': value_type}
						leaf._transform()
						leaf.value = live_value
						leaf.init_value = init_value
						pos.opc_children.append(attr_name)
						setattr(pos, attr_name, leaf)
					if 'CRValues' in file_path:
						leaf.idx_prop = {5002:'ColdRetain'}
					elif 'RValues' in file_path:
						leaf.idx_prop = {5002: 'Retain'}
	return root


if __name__ == "__main__":
	lib_dict = {'Real':'opc_vars','Bool':'opc_vars','Time':'opc_vars','Uint':'opc_vars','Dint':'opc_vars','Dword':'opc_vars','Word':'opc_vars','Date_And_Time':'opc_vars','String':'opc_vars'}
	lib_and_file_to_create = OrderedDict()
	lib_and_file_to_create['opc_class_lib.System']  = 'input\\System.xlsx'
	lib_and_file_to_create['opc_class_lib.BasicLib_1_8_5'] = 'input\\BasicLib_1_8_5.xlsx'
	for new_lib_name, from_file in lib_and_file_to_create.items():
		wb = load_workbook(filename=from_file, read_only=True)
		output, lib_dict = make_lib_file(wb,new_lib_name,lib_dict)
		with open(new_lib_name.replace('.','\\') +".py",'w', encoding='utf-8') as outputFile:
			outputFile.write(output)
	update_init_file()
