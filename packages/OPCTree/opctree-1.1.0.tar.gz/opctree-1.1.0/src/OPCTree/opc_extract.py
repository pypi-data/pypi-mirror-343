#import settings
from . import opc_fetch, struct_identi
from openpyxl import Workbook



variables = opc_fetch.connect_and_extract_variables()

wb = Workbook()
ws = wb.active

startRow = 1
startCol = 1

for colNbr in range(len(variables[1])):
	cell = ws.cell(row=1, column=colNbr+startCol, value=variables[1][colNbr][-2])
for rowNbr in range(len(variables)):
	for colNbr in range(len(variables[rowNbr])):
		cell = ws.cell(row=rowNbr+startRow +1, column=colNbr+startCol, value=variables[rowNbr][colNbr][-1])
				
wb.save("Output//FromOPC.xlsx")

known_structs = struct_identi.get_structure(variables)