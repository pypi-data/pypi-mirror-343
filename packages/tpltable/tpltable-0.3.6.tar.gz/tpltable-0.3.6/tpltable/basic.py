from typing import Union, Union as U, List, Tuple, Dict, Any, Optional
from copy import copy
import pandas as pd
import numpy as np
import warnings
import colorama
import random
import time
import re

import openpyxl as yxl
from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.cell.cell import Cell
from openpyxl.utils import get_column_letter, column_index_from_string as _column_index_from_string
from openpyxl.utils.cell import coordinate_from_string as _coordinate_from_string
from openpyxl.styles import Font, Color, Alignment, Border, Side, PatternFill

NaN = pd.NA
nan = pd.NA

def isnan(value):
    return pd.isna(value) or str(value) == 'nan'

# NOTE: Colorama Start -------------------------------------------------------------------------->
# 初始化colorama
colorama.init(autoreset=True)

# 定义颜色和样式映射
COLORAMA_COLORS = {
    'red': colorama.Fore.RED,
    'green': colorama.Fore.GREEN,
    'yellow': colorama.Fore.YELLOW,
    'blue': colorama.Fore.BLUE,
    'magenta': colorama.Fore.MAGENTA,
    'cyan': colorama.Fore.CYAN,
    'white': colorama.Fore.WHITE,
    'bright_black': colorama.Fore.BLACK + colorama.Style.BRIGHT,
    'bright_red': colorama.Fore.RED + colorama.Style.BRIGHT,
    'bright_green': colorama.Fore.GREEN + colorama.Style.BRIGHT,
    'bright_yellow': colorama.Fore.YELLOW + colorama.Style.BRIGHT,
    'bright_blue': colorama.Fore.BLUE + colorama.Style.BRIGHT,
    'bright_magenta': colorama.Fore.MAGENTA + colorama.Style.BRIGHT,
    'bright_cyan': colorama.Fore.CYAN + colorama.Style.BRIGHT,
    'bright_white': colorama.Fore.WHITE + colorama.Style.BRIGHT,
}

COLORAMA_BG_COLORS = {
    'red': colorama.Back.RED,
    'green': colorama.Back.GREEN,
    'yellow': colorama.Back.YELLOW,
    'blue': colorama.Back.BLUE,
    'magenta': colorama.Back.MAGENTA,
    'cyan': colorama.Back.CYAN,
    'white': colorama.Back.WHITE,
    'bright_black': colorama.Back.BLACK + colorama.Style.BRIGHT,
    'bright_red': colorama.Back.RED + colorama.Style.BRIGHT,
    'bright_green': colorama.Back.GREEN + colorama.Style.BRIGHT,
    'bright_yellow': colorama.Back.YELLOW + colorama.Style.BRIGHT,
    'bright_blue': colorama.Back.BLUE + colorama.Style.BRIGHT,
    'bright_magenta': colorama.Back.MAGENTA + colorama.Style.BRIGHT,
    'bright_cyan': colorama.Back.CYAN + colorama.Style.BRIGHT,
    'bright_white': colorama.Back.WHITE + colorama.Style.BRIGHT,
}

# 定义样式
COLORAMA_STYLES = {
    'dim': colorama.Style.DIM,
    'normal': colorama.Style.NORMAL,
    'bright': colorama.Style.BRIGHT,
}
# NOTE: Colorama Ready --------------------------------------------------------------------------;

__PAT_COORD = re.compile("([A-Z]+)(\d+)")


# 模糊的行列索引混用警告
class AmbiguousIndexWarning(UserWarning):
    ...

class Coord:
    def __init__(self, iloc_c, iloc_r):
        self._r:int = iloc_r
        self._c:int = iloc_c

    @property
    def r(self):
        return str(self._r + 1)

    @r.setter
    def r(self, digit:str):
        self._r = int(digit) - 1

    @property
    def c(self):
        return get_column_letter(self._c + 1)

    @c.setter
    def c(self, letter:str):
        self._c = _column_index_from_string(letter) - 1

    @property
    def ir(self):
        return self._r

    @ir.setter
    def ir(self, value:int):
        self._r = value

    @property
    def ic(self):
        return self._c

    @ic.setter
    def ic(self, value:int):
        self._c = value

    @property
    def letter(self):
        return f"{self.c}{self.r}"

    def __str__(self):
        return f"{self.c}{self.r}|slice{(self._c, self._r)}"

    def __repr__(self):
        return f"Coord({self.c}{self.r})"

    def copy(self):
        return Coord(self._c, self._r)

    @classmethod
    def FromLetter(cls, letter:str):
        _ = index_from_string(letter)
        if isinstance(_, Coord):
            return _
        return Coord(_[0] - 1, _[1] - 1)


def index_from_string(letters:str, *, only_int:bool=False, should:str=None) -> Union[int, Coord]:
    """
    Convert a column letter into a number.
    :param letters: str. The letter to convert.
    :param only_int: bool. If True, will throw if the letter is a Coord.  [Error]
    :param should: str. The type of the letter should be. Default is None. [Warning]
    :return: int| (int, int). The number or coord of string.
    # 除了Coord外，其他的都是从1开始的
    """
    if letters.isdigit():
        if should == 's':
            warnings.warn(
                AmbiguousIndexWarning(f"Here should be a letter (Col-Index), but got '{letters}' (Row-Index). Regarded as '{get_column_letter(int(letters))}'."),
            )
        return int(letters)
    _s = __PAT_COORD.match(letters)
    if _s:
        if only_int:
            raise ValueError(f"Here only accept int, but got '{letters}'")
        _c, _r = _s.group(1), _s.group(2)
        return Coord(index_from_string(_c) - 1, int(_r) - 1)
    if should == 'i':
        warnings.warn(
            AmbiguousIndexWarning(f"Here should be an digit (Row-Index), but got '{letters}' (Col-Index). Regarded as '{_column_index_from_string(letters)}'.")
        )
    return _column_index_from_string(letters)


def split_range_from_string(range_letter:str) -> Tuple[str, str]:
    """
    Convert a range letter into a number.
    * do not have ':' will down to 'index_from_string'.
    * have multiple ':' will cause an Error.
    :param range_letter: str. The column letter to convert.
    :return: tuple of raw label
    """
    _colon_times = range_letter.count(':')
    assert _colon_times == 1, f"Unexcepted range-index: '{range_letter}'.(Expected one ':')"

    _colon_index = range_letter.find(":")
    _start_letter, _end_letter = range_letter[:_colon_index], range_letter[_colon_index + 1:]

    # Check if the two are in the same type
    if _start_letter.isdigit() != _end_letter.isdigit():  # Can not be both Coord & same type
        raise ValueError(f"{range_letter} is not in the same type(not a valid range)")

    return _start_letter, _end_letter

def reinsert_indent(text, keep_raw:bool=False, *, indent:str=None, strip:str=None) -> str:
    """
    Reinsert the indent of each line.
    :param text: str. The text to reinsert the indent.
    :param keep_raw: bool. Keep the raw indent of the each line. / If False, strip each line
    *
    :param indent: str. The indent to reinsert. Default is \t.
    :param strip: str. The characters to strip from the start of each line. Default is " \t".
    :return: str. The text with the indent reinserted.
    """
    if indent is None:
        indent = "\t"
    if strip is None:
        strip = " \t"

    lines = [line for line in text.split("\n") if line]
    if not lines:
        return text

    if not keep_raw:
        lines = [line.lstrip(strip) for line in lines]

    return "\n".join(indent + line for line in lines)


def integer_slice(slice_obj, *, only_1d:bool=False, only_int:bool=False) -> slice:
    """
    Make sure the slice object is an integer slice. Will Auto convert the letter to number.
    :param slice_obj: slice. The slice object to convert.
    :param only_1d: bool. If True, will throw if the slice is not 1-d.
    :param only_int: bool. If True, will throw if one of the start or stop is not int.
    :return: slice. The converted slice object.
    """
    start, stop, step = slice_obj.start, slice_obj.stop, slice_obj.step
    if isinstance(start, str):
        start = index_from_string(start) - 1
    if isinstance(stop, str):
        stop = index_from_string(stop) - 1
    if only_1d:
        assert not isinstance(start, Coord) and not isinstance(stop, Coord), f"Here only accept 1-d slice, but got {start} and {stop}"
    if only_int:
        assert isinstance(start, int) and isinstance(stop, int), f"Here only accept int, but got {start} and {stop}"
    return slice(start, stop, step)





def color_string(txt, color=None, bg_color=None, style=None):
    """
    Color the text with optional background color and style.
    * won't auto use black color.
    :param txt: str. The text to color.
    :param color: str. The foreground color to use. Default is None, which means random color.
    :param bg_color: str. The background color to use. Default is None, which means no background color.
    :param style: str. The style to apply. Default is None, which means no style.
    :return: str. The colored text.
    """

    if color is None:
        color = random.choice(list(COLORAMA_COLORS.keys()))

    # 构建颜色字符串
    color_code = COLORAMA_COLORS.get(color.lower())
    on_color_code = COLORAMA_BG_COLORS.get(bg_color.lower()) if bg_color else ""
    style_code = COLORAMA_STYLES.get(style.lower()) if style else ""

    # 应用颜色和样式
    colored_text = ""
    if on_color_code:
        colored_text += on_color_code
    if color_code:
        colored_text += color_code
    if style_code:
        colored_text += style_code

    colored_text += txt + colorama.Style.RESET_ALL

    return colored_text

class TimeRecorder:
    def __init__(self, start:bool=True):
        self._start = None
        self._end = None

        if start:
            self.start()

    def start(self):
        self._start = time.time()
        self._end = None

    def end(self):
        self._end = time.time()

    def duration(self):
        if self._end is None:
            return time.time() - self._start
        if self._start is None:
            return -1
        return self._end - self._start

    def __str__(self):
        return f"TimeRecorder({self.duration()}s)"

    def __repr__(self):
        return f"{self.duration()}s)"

    def tick(self) -> float:
        """
        End the current time and start a new one.
        :return: float. The duration of the last time.
        """
        self.end()
        _ = self.duration()
        self.start()
        return _

    @property
    def s(self):
        return round(self.duration(), 5)

    @property
    def ms(self):
        return round(self.duration() * 1000, 2)

    def ds(self, div:int=1, r:int=2):
        return round(self.s / div, r)

    def dms(self, div:int=1, r:int=2):
        return round(self.ms / div, r)

if __name__ == '__main__':
    print(
        _column_index_from_string('A')
    )
    exit()
    print(
        index_from_string("A"),
        index_from_string("1"),
        index_from_string("1", should='s'),
        index_from_string("A", should='i'),
        index_from_string("A1"),
    )
