import openpyxl as xl
from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.cell.cell import Cell
from openpyxl.styles import Font, Color, Alignment, Border, Side, PatternFill
from openpyxl.styles.colors import BLACK, WHITE
from openpyxl.utils import get_column_letter
from copy import copy

# 定义样式类型常量
TYPE_COLOR = "t_color"  # 前景色
TYPE_BCOLOR = "t_bcolor"  # 背景色
TYPE_FONT_TYPE = "t_font_type"  # 字体
TYPE_FONT_SIZE = "t_font_size"  # 字号
TYPE_FONT_BOLD = "t_font_bold"  # 粗体
TYPE_FONT_ITALIC = "t_font_italic"  # 斜体
TYPE_UNDERLINE = "t_underline"  # 下划线
TYPE_DELETE_LINE = "t_delete_line"  # 删除线
TYPE_HALIGN = "t_halign"  # 水平对齐
TYPE_VALIGN = "t_valign"  # 垂直对齐
TYPE_BORDER = "t_border"  # 边框
TYPE_ROW_HEIGHT = "t_row_height"  # 行高
TYPE_COL_WIDTH = "t_col_width"  # 列宽
TYPE_DATA_FORMAT = "t_data_format"  # 数据格式
TYPE_AUTO_NEWLINE = "t_auto_newline"  # 自动换行
F_STRING = '@'  # 文本格式
F_NUMBER = '0'  # 数字格式，无小数点
F_DATE = 'YYYY-MM-DD'  # 日期格式，示例为年-月-日
F_TIME = 'HH:MM:SS'  # 时间格式，示例为时:分:秒
...


class TCell:
    def __init__(self, wb: Workbook, ws: Worksheet, cell: Cell):
        self.wb = wb
        self.ws = ws
        self.cell = cell


def SetCellStyle(tc: TCell, style_type: str, value):
    # 先保存当前的样式，以便我们可以修改并重新应用
    current_font = tc.cell.font
    current_alignment = tc.cell.alignment
    new_font = copy(current_font)  # 使用copy来创建一个可修改的副本
    new_alignment = copy(current_alignment)  # 使用copy来创建一个可修改的副本

    if style_type == TYPE_COLOR:
        new_font.color = Color(value)
    elif style_type == TYPE_BCOLOR:
        tc.cell.fill = PatternFill(fgColor=value, bgColor=value, fill_type='solid')
    elif style_type == TYPE_FONT_TYPE:
        new_font.name = value
    elif style_type == TYPE_FONT_SIZE:
        new_font.size = value
    elif style_type == TYPE_FONT_BOLD:
        new_font.bold = value
    elif style_type == TYPE_FONT_ITALIC:
        new_font.italic = value
    elif style_type == TYPE_UNDERLINE:
        if value:
            new_font.underline = 'single'
        else:
            new_font.underline = None
    elif style_type == TYPE_DELETE_LINE:
        new_font.strikethrough = value
    elif style_type == TYPE_HALIGN:
        new_alignment.horizontal = value
    elif style_type == TYPE_VALIGN:
        new_alignment.vertical = value
    elif style_type == TYPE_AUTO_NEWLINE:
        new_alignment.wrapText = value
    elif style_type == TYPE_BORDER:
        tc.cell.border = value
    if style_type == TYPE_ROW_HEIGHT:
        tc.ws.row_dimensions[tc.cell.row].height = value  # 设置行高
    elif style_type == TYPE_COL_WIDTH:
        # 将数字索引转换为列字母标识
        col_letter = get_column_letter(tc.cell.column)
        tc.ws.column_dimensions[col_letter].width = value  # 设置列宽
    elif style_type == TYPE_DATA_FORMAT:
        tc.cell.number_format = value  # 设置数据格式

    # 重新应用修改后的样式
    tc.cell.font = new_font
    tc.cell.alignment = new_alignment


def GetCellStyle(tc: TCell, style_type: str) -> object:
    if style_type == TYPE_COLOR:
        return tc.cell.font.color
    elif style_type == TYPE_BCOLOR:
        return tc.cell.fill.fgColor
    elif style_type == TYPE_FONT_TYPE:
        return tc.cell.font.name
    elif style_type == TYPE_FONT_SIZE:
        return tc.cell.font.size
    elif style_type == TYPE_FONT_BOLD:
        return tc.cell.font.bold
    elif style_type == TYPE_FONT_ITALIC:
        return tc.cell.font.italic
    elif style_type == TYPE_UNDERLINE:
        return tc.cell.font.underline == 'single'
    elif style_type == TYPE_DELETE_LINE:
        return tc.cell.font.strikethrough
    elif style_type == TYPE_HALIGN:
        return tc.cell.alignment.horizontal
    elif style_type == TYPE_VALIGN:
        return tc.cell.alignment.vertical
    elif style_type == TYPE_AUTO_NEWLINE:
        return tc.cell.alignment.wrapText
    elif style_type == TYPE_BORDER:
        return tc.cell.border
    elif style_type == TYPE_ROW_HEIGHT:
        return tc.ws.row_dimensions[tc.cell.row].height  # 获取行高(Strange, why not tc.cell.row_dimensions is same as tc.cell.column_dimensions? One is str, the other is int.)
    elif style_type == TYPE_COL_WIDTH:
        col_letter = get_column_letter(tc.cell.column)
        return tc.ws.column_dimensions[col_letter].width  # 获取列宽
    elif style_type == TYPE_DATA_FORMAT:
        return tc.cell.number_format
    return None



def _Style_Op_Demo():
    # 示例使用
    wb = xl.Workbook()
    ws = wb.active
    cell = ws['A1']
    cell.value = 'Hello, World!'
    tcell = TCell(wb, ws, cell)

    # 设置样式
    SetCellStyle(tcell, TYPE_COLOR, 'FF0000')  # 设置红色前景色
    SetCellStyle(tcell, TYPE_BCOLOR, '00FF00')  # 设置绿色背景色
    SetCellStyle(tcell, TYPE_FONT_TYPE, 'Arial')  # 设置字体
    SetCellStyle(tcell, TYPE_FONT_SIZE, 20)  # 设置字号
    SetCellStyle(tcell, TYPE_FONT_BOLD, True)  # 设置粗体
    SetCellStyle(tcell, TYPE_FONT_ITALIC, True)  # 设置斜体
    SetCellStyle(tcell, TYPE_UNDERLINE, True)  # 设置下划线
    SetCellStyle(tcell, TYPE_DELETE_LINE, True)  # 设置删除线
    SetCellStyle(tcell, TYPE_HALIGN, 'center')  # 设置水平居中
    SetCellStyle(tcell, TYPE_VALIGN, 'center')  # 设置垂直居中
    SetCellStyle(tcell, TYPE_BORDER, Border(left=Side(border_style='thin', color=BLACK),
                                            right=Side(border_style='thin', color=BLACK),
                                            top=Side(border_style='thin', color=BLACK),
                                            bottom=Side(border_style='thin', color=BLACK)))  # 设置边框
    SetCellStyle(tcell, TYPE_ROW_HEIGHT, 40)  # 设置行高
    SetCellStyle(tcell, TYPE_COL_WIDTH, 100)  # 设置列宽
    SetCellStyle(tcell, TYPE_DATA_FORMAT, F_STRING)  # 设置数据格式

    # 获取样式
    print("fg", GetCellStyle(tcell, TYPE_COLOR), "\n")  # 获取前景色
    print("bg", GetCellStyle(tcell, TYPE_BCOLOR), "\n")  # 获取前景色
    print("bold", GetCellStyle(tcell, TYPE_FONT_BOLD), "\n")  # 获取是否为粗体
    print("ita", GetCellStyle(tcell, TYPE_FONT_ITALIC), "\n")  # 获取是否为斜体
    print("under", GetCellStyle(tcell, TYPE_UNDERLINE), "\n")  # 获取是否有下划线
    print("del", GetCellStyle(tcell, TYPE_DELETE_LINE), "\n")  # 获取是否有删除线
    print("ha", GetCellStyle(tcell, TYPE_HALIGN), "\n")  # 获取水平对齐方式
    print("va", GetCellStyle(tcell, TYPE_VALIGN), "\n")  # 获取垂直对齐方式
    bd = GetCellStyle(tcell, TYPE_BORDER)
    print("border", bd, "\n")  # 获取边框
    print("row", GetCellStyle(tcell, TYPE_ROW_HEIGHT), "\n")  # 获取行高
    print("col", GetCellStyle(tcell, TYPE_COL_WIDTH), "\n")  # 获取列宽
    print("format", GetCellStyle(tcell, TYPE_DATA_FORMAT), "\n")  # 获取数据格式

    # B1
    cell = ws['B1']
    cell.value = 114514
    tcell = TCell(wb, ws, cell)
    SetCellStyle(tcell, TYPE_DATA_FORMAT, '0.00')  # 设置数据格式
    SetCellStyle(tcell, TYPE_COL_WIDTH, 30)  # 设置列宽

    # A2 a Data
    cell = ws['A2']
    cell.value = '2021-07-01'
    tcell = TCell(wb, ws, cell)
    SetCellStyle(tcell, TYPE_DATA_FORMAT, 'YYYY-MM-DD')  # 设置数据格式
    # 居中
    SetCellStyle(tcell, TYPE_HALIGN, 'center')  # 设置水平居中
    SetCellStyle(tcell, TYPE_VALIGN, 'center')  # 设置垂直居中

    # B2 is B1 ** 2
    cell = ws['B2']
    cell.value = '=B1^2'
    tcell = TCell(wb, ws, cell)
    SetCellStyle(tcell, TYPE_DATA_FORMAT, '0.00')  # 设置数据格式

    # D1
    cell = ws['D1']
    cell.value = 'Hello, World!Hello, World!'
    tcell = TCell(wb, ws, cell)
    SetCellStyle(tcell, TYPE_AUTO_NEWLINE, True)  # 设置自动换行

    wb.save('test.xlsx')

if __name__ == '__main__':
    _Style_Op_Demo()