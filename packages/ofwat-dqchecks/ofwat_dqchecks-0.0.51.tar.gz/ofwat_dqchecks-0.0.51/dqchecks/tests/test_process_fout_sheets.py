"""
Tests for the transforms.py file
"""
from datetime import datetime
from unittest.mock import patch
import pytest
from openpyxl.workbook.workbook import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import pandas as pd
from dqchecks.transforms import (
    process_fout_sheets,
    extract_fout_sheets,
    ProcessingContext)

@pytest.fixture
def valid_context():
    """
    Fixture for valid ProcessingContext
    """
    return ProcessingContext(
        org_cd="ORG001",
        submission_period_cd="2025Q1",
        process_cd="PROCESS01",
        template_version="1.0",
        last_modified=datetime(2025, 3, 3)
    )

@pytest.fixture
def empty_workbook_without_fout():
    """
    Fixture for an empty workbook
    """
    wb = Workbook()
    wb.create_sheet("other")
    return wb

@pytest.fixture
def workbook_with_data():
    """
    Fixture for workbook with data in 'fOut_*' sheets
    """
    wb = Workbook()
    sheet = wb.create_sheet("fOut_Sheet1")
    sheet.append(["", "", "", "", "", "", "", ""])
    sheet.append(["Acronym", "Reference", "Item description", "Unit", "Model",
                  "Description_input", "Constant", "2020-21"])
    sheet.append(["", "", "", "", "", "", "", ""])
    sheet.append(["a", "a", "a", "a", "a", "a", "a", "a"])
    sheet.append(["a", "a", "a", "a", "a", "a", "a", "a"])
    return wb

# pylint: disable=W0621
def test_process_fout_sheets_valid(workbook_with_data, valid_context):
    """
    Test valid processing of a workbook
    """
    observation_patterns = [r'^\s*2[0-9]{3}-[1-9][0-9]\s*$']
    fout_patterns = ["^fOut_"]
    result_df = process_fout_sheets(
        workbook_with_data,
        valid_context,
        observation_patterns,
        fout_patterns)
    assert isinstance(result_df, pd.DataFrame)
    assert not result_df.empty

    expected_columns = [
        "Organisation_Cd",
        "Submission_Period_Cd",
        "Observation_Period_Cd",
        "Process_Cd",
        "Template_Version",
        "Sheet_Cd",
        "Measure_Cd",
        "Measure_Value",
        "Measure_Unit",
        "Model_Cd",
        "Submission_Date",
        "Section_Cd",
        "Cell_Cd",
    ]
    assert all(i in result_df.columns for i in expected_columns)
    # assert "Organisation_Cd" in result_df.columns
    # assert "Observation_Period_Cd" in result_df.columns
    assert result_df["Sheet_Cd"].iloc[0] == "fOut_Sheet1"

    # Patch the logging to capture the warning message
    with patch("logging.warning") as mock_warning:
        # Call the function (this will trigger the warning if wb.data_only is False)
        process_fout_sheets(
            workbook_with_data,
            valid_context,
            observation_patterns,
            fout_patterns)

        # Check that the warning was logged with the expected message
        mock_warning.assert_called_with(
            "Reading in non data_only mode. Some data may not be accessible.")

# pylint: disable=W0621
def test_process_fout_sheets_invalid_workbook(valid_context):
    """
    Test invalid workbook type (not an openpyxl Workbook)
    """
    observation_patterns = [r'^\s*2[0-9]{3}-[1-9][0-9]\s*$']
    fout_patterns = ["^fOut_"]
    with pytest.raises(TypeError,
            match="The 'wb' argument must be a valid openpyxl workbook object."):
        process_fout_sheets(
            "invalid",
            valid_context,
            observation_patterns,
            fout_patterns)

# pylint: disable=W0621
def test_process_fout_sheets_no_fout_sheets(empty_workbook_without_fout, valid_context):
    """
    Test missing 'fOut_*' sheets in the workbook
    """
    observation_patterns = [r'^\s*2[0-9]{3}-[1-9][0-9]\s*$']
    fout_patterns = ["^fOut_"]
    with pytest.raises(ValueError, match="No sheets matching patterns"):
        process_fout_sheets(
            empty_workbook_without_fout,
            valid_context,
            observation_patterns,
            fout_patterns)

# pylint: disable=W0621
def test_process_fout_sheets_missing_observation_columns(workbook_with_data, valid_context):
    """
    Test missing observation period columns (no 'yyyy-yy' pattern columns)
    """
    observation_patterns = [r'^\s*2[0-9]{3}-[1-9][0-9]\s*$']
    sheet = workbook_with_data["fOut_Sheet1"]
    fout_patterns = ["^fOut_"]
    # Remove the observation period columns to simulate the case
    sheet.delete_cols(7, 2)

    with pytest.raises(ValueError, match="No observation period columns found in the data."):
        process_fout_sheets(
            workbook_with_data,
            valid_context,
            observation_patterns,
            fout_patterns)

# pylint: disable=W0621
def test_process_fout_sheets_drop_nan_rows(workbook_with_data, valid_context):
    """
    Test that dropping NaN rows works as expected
    """
    observation_patterns = [r'^\s*2[0-9]{3}-[1-9][0-9]\s*$']
    fout_patterns = ["^fOut_"]
    # Create a sheet with NaN rows
    sheet = workbook_with_data["fOut_Sheet1"]
    sheet.append([None, None, None, None])  # Add a row with all NaNs

    result_df = process_fout_sheets(
        workbook_with_data,
        valid_context,
        observation_patterns,
        fout_patterns)
    assert result_df.shape[0] == 3  # Should have dropped the NaN row
    assert result_df["Sheet_Cd"].iloc[0] == "fOut_Sheet1"

# pylint: disable=W0621
def test_process_fout_sheets_invalid_context_org_cd(workbook_with_data):
    """
    Test when context has invalid 'org_cd' (empty string)
    """
    observation_patterns = [r'^\s*2[0-9]{3}-[1-9][0-9]\s*$']
    fout_patterns = ["^fOut_"]
    invalid_context = ProcessingContext(
        org_cd="",
        submission_period_cd="2025Q1",
        process_cd="PROCESS01",
        template_version="1.0",
        last_modified=datetime(2025, 3, 3)
    )
    with pytest.raises(ValueError, match="The 'org_cd' argument must be a non-empty string."):
        process_fout_sheets(
            workbook_with_data,
            invalid_context,
            observation_patterns,
            fout_patterns)

# pylint: disable=W0621
def test_process_fout_sheets_invalid_context_submission_period(workbook_with_data):
    """
    Test when context has invalid 'submission_period_cd' (None)
    """
    observation_patterns = [r'^\s*2[0-9]{3}-[1-9][0-9]\s*$']
    fout_patterns = ["^fOut_"]
    invalid_context = ProcessingContext(
        org_cd="ORG001",
        submission_period_cd=None,
        process_cd="PROCESS01",
        template_version="1.0",
        last_modified=datetime(2025, 3, 3)
    )
    with pytest.raises(ValueError,
            match="The 'submission_period_cd' argument must be a non-empty string."):
        process_fout_sheets(
            workbook_with_data,
            invalid_context,
            observation_patterns,
            fout_patterns)

# pylint: disable=W0621
def test_process_fout_sheets_invalid_context_last_modified(workbook_with_data):
    """
    Test when context has invalid 'last_modified' (wrong type)
    """
    observation_patterns = [r'^\s*2[0-9]{3}-[1-9][0-9]\s*$']
    fout_patterns = ["^fOut_"]
    invalid_context = ProcessingContext(
        org_cd="ORG001",
        submission_period_cd="2025Q1",
        process_cd="PROCESS01",
        template_version="1.0",
        last_modified="invalid"  # Should be a datetime object
    )
    with pytest.raises(ValueError, match="The 'last_modified' argument must be a datetime object."):
        process_fout_sheets(
            workbook_with_data,
            invalid_context,
            observation_patterns,
            fout_patterns)

# pylint: disable=W0621
def test_process_fout_sheets_empty_data(workbook_with_data, valid_context):
    """
    Test if the function raises error when no valid rows are available after dropping NaNs
    """
    observation_patterns = [r'^\s*2[0-9]{3}-[1-9][0-9]\s*$']
    fout_patterns = ["^fOut_"]
    sheet = workbook_with_data["fOut_Sheet1"]
    sheet.delete_rows(0, 5)  # Remove all data rows

    with pytest.raises(ValueError, match="Sheet 'fOut_Sheet1' is empty or has no data."):
        process_fout_sheets(
            workbook_with_data,
            valid_context,
            observation_patterns,
            fout_patterns)

def convert_dataframe_to_rows(df, index=True, header=True):
    """
    Helper function to convert a pandas DataFrame to rows for openpyxl.
    """
    return dataframe_to_rows(df, index=index, header=header)

def create_openpyxl_workbook(sheet_data):
    """
    Helper function to create an openpyxl Workbook from a dictionary of DataFrames.
    Each key in the dictionary represents a sheet name and each value is a DataFrame
    for that sheet.
    """
    wb = Workbook()

    # Remove the default sheet created
    wb.remove(wb.active)

    for sheet_name, df in sheet_data.items():
        ws = wb.create_sheet(sheet_name)
        # Write the DataFrame to the sheet
        for row in convert_dataframe_to_rows(df, index=False, header=True):
            ws.append(row)

    return wb

def test_different_observation_periods():
    """
    Test case to check if the function handles cases where different 
    sheets have different observation periods.
    
    This test verifies that the function correctly handles sheets with
    different sets of observation periods, ensuring that each sheet's
    observation periods are melted properly without causing errors.
    """
    # Create data for two sheets, each with different observation periods
    sheet_data = {
        'fOut_2023': pd.DataFrame({
            "Reference": ["Reference", "", "Ref1", "Ref2", "Ref3"],
            "Item description": ["Item description", "", "Item 1", "Item 2", "Item 3"],
            "Unit": ["Unit", "", "kg", "g", "lbs"],
            "Model": ["Model", "", "A", "B", "C"],
            "2020-21": ["2020-21", "", 10, 20, 30],
            "2021-22": ["2021-22", "", 15, 25, 35],
        }),
        'fOut_2024': pd.DataFrame({
            "Reference": ["Reference", "", "Ref1", "Ref2", "Ref3"],
            "Item description": ["Item description", "", "Item 1", "Item 2", "Item 3"],
            "Unit": ["Unit", "", "kg", "g", "lbs"],
            "Model": ["Model", "", "A", "B", "C"],
            "2021-22": ["2021-22", "", 10, 20, 30],
            "2022-23": ["2022-23", "", 15, 25, 35],
        }),
    }

    # Create an openpyxl workbook from the sheet data
    wb = create_openpyxl_workbook(sheet_data)

    observation_patterns = [r'^\s*2[0-9]{3}-[1-9][0-9]\s*$']
    fout_patterns = ["^fOut_"]

    context = ProcessingContext(
        org_cd="ORG123",
        submission_period_cd="2025Q1",
        process_cd="process_1",
        template_version="v1.0",
        last_modified=datetime(2025, 2, 11),
    )

    # Process the file and get the result DataFrame
    result_df = process_fout_sheets(
        wb,
        context,
        observation_patterns,
        fout_patterns)

    # Check that all observation periods from both sheets are present
    expected_observation_periods = {"2020-21", "2021-22", "2022-23"}
    assert set(result_df["Observation_Period_Cd"]) == expected_observation_periods

    # Check observation periods for each sheet
    assert set(
        result_df[result_df["Sheet_Cd"] == "fOut_2023"]["Observation_Period_Cd"]
        ) == {"2020-21", "2021-22"}
    assert set(
        result_df[result_df["Sheet_Cd"] == "fOut_2024"]["Observation_Period_Cd"]
        ) == {"2021-22", "2022-23"}

    # Check that the resulting DataFrame is not empty
    assert not result_df.empty

# Test when context has invalid 'process_cd' (non-string or empty string)
# pylint: disable=W0621
def test_process_fout_sheets_invalid_context_process_cd(workbook_with_data):
    """
    Test case to check if the function raises an error when the 'process_cd' context is invalid.
    """
    # Test case where process_cd is an empty string
    invalid_context = ProcessingContext(
        org_cd="ORG001",
        submission_period_cd="2025Q1",
        process_cd="",  # Invalid: empty string
        template_version="1.0",
        last_modified=datetime(2025, 3, 3)
    )
    observation_patterns = [r'^\s*2[0-9]{3}-[1-9][0-9]\s*$']
    fout_patterns = ["^fOut_"]

    with pytest.raises(ValueError, match="The 'process_cd' argument must be a non-empty string."):
        process_fout_sheets(
            workbook_with_data,
            invalid_context,
            observation_patterns,
            fout_patterns)

    # Test case where process_cd is not a string (e.g., an integer)
    invalid_context = ProcessingContext(
        org_cd="ORG001",
        submission_period_cd="2025Q1",
        process_cd=1234,  # Invalid: integer
        template_version="1.0",
        last_modified=datetime(2025, 3, 3)
    )

    with pytest.raises(ValueError, match="The 'process_cd' argument must be a non-empty string."):
        process_fout_sheets(
            workbook_with_data,
            invalid_context,
            observation_patterns,
            fout_patterns)


# Test when context has invalid 'template_version' (non-string or empty string)
# pylint: disable=W0621
def test_process_fout_sheets_invalid_context_template_version(workbook_with_data):
    """
    Test case to check if the function raises an error
    when the 'template_version' context is invalid.
    """
    # Test case where template_version is an empty string
    invalid_context = ProcessingContext(
        org_cd="ORG001",
        submission_period_cd="2025Q1",
        process_cd="PROCESS01",
        template_version="",  # Invalid: empty string
        last_modified=datetime(2025, 3, 3)
    )
    observation_patterns = [r'^\s*2[0-9]{3}-[1-9][0-9]\s*$']
    fout_patterns = ["^fOut_"]

    with pytest.raises(ValueError,
            match="The 'template_version' argument must be a non-empty string."):
        process_fout_sheets(
            workbook_with_data,
            invalid_context,
            observation_patterns,
            fout_patterns)

    # Test case where template_version is not a string (e.g., a number)
    invalid_context = ProcessingContext(
        org_cd="ORG001",
        submission_period_cd="2025Q1",
        process_cd="PROCESS01",
        template_version=1.0,  # Invalid: not a string
        last_modified=datetime(2025, 3, 3)
    )

    with pytest.raises(ValueError,
            match="The 'template_version' argument must be a non-empty string."):
        process_fout_sheets(
            workbook_with_data,
            invalid_context,
            observation_patterns,
            fout_patterns)

# Test when observation_patterns is not a list
# pylint: disable=W0621
def test_process_fout_sheets_invalid_observation_patterns_not_list(
        workbook_with_data, valid_context):
    """
    Test case to check if the function raises an error when 'observation_patterns' is not a list.
    """
    invalid_observation_patterns = "invalid_pattern_string"  # Not a list
    fout_patterns = ["^fOut_"]

    with pytest.raises(ValueError,
            match="The 'observation_patterns' argument needs to be a list of regex strings."):
        process_fout_sheets(
            workbook_with_data,
            valid_context,
            invalid_observation_patterns,
            fout_patterns)

# Test when observation_patterns contains elements that are not strings
# pylint: disable=W0621
def test_process_fout_sheets_invalid_observation_patterns_non_string(
        workbook_with_data, valid_context):
    """
    Test case to check if the function raises an error when
    'observation_patterns' contains elements that are not strings.
    """
    # List with non-string element (integer)
    invalid_observation_patterns = [r'^\s*2[0-9]{3}-[1-9][0-9]\s*$', 1234]
    fout_patterns = ["^fOut_"]

    with pytest.raises(ValueError,
            match="The 'observation_patterns' argument needs to be a list of regex strings."):
        process_fout_sheets(
            workbook_with_data,
            valid_context,
            invalid_observation_patterns,
            fout_patterns)

# Test when observation_patterns contains invalid regex patterns
# pylint: disable=W0621
def test_process_fout_sheets_invalid_observation_patterns_invalid_regex(
        workbook_with_data, valid_context):
    """
    Test case to check if the function raises an error
    when 'observation_patterns' contains invalid regex patterns.
    """
    invalid_observation_patterns = [r'^\s*2[0-9]{3}-[1-9][0-9]\s*$', r"[^"]
    fout_patterns = ["^fOut_"]

    with pytest.raises(ValueError,
            match="The 'observation_patterns' argument needs to be a list of regex strings."):
        process_fout_sheets(
            workbook_with_data,
            valid_context,
            invalid_observation_patterns,
            fout_patterns)

@pytest.fixture
def workbook_with_invalid_data():
    """
    Fixture for workbook with no valid rows after dropping NaNs
    """
    wb = Workbook()
    sheet = wb.create_sheet("fOut_Sheet1")
    sheet.append(["", "", "", "", "", "", "", ""])  # Blank row
    sheet.append([None, None, None, None, None, None, None, None])  # Another row with all NaNs
    return wb

# pylint: disable=W0621
def test_process_fout_sheets_no_valid_rows(workbook_with_invalid_data, valid_context):
    """
    Test case to check that an error is raised when no valid rows remain after \
        dropping NaN rows
    """
    observation_patterns = [r'^\s*2[0-9]{3}-[1-9][0-9]\s*$']
    fout_patterns = ["^fOut_"]

    # Process the sheet
    with pytest.raises(ValueError,
            match="No valid data found after removing rows with NaN values."):
        process_fout_sheets(
            workbook_with_invalid_data,
            valid_context,
            observation_patterns,
            fout_patterns)

def create_workbook_with_sheets(sheet_names):
    """Helper function to create workbook with sheets"""
    wb = Workbook()
    # Remove the default sheet created by openpyxl
    default_sheet = wb.active
    wb.remove(default_sheet)

    for name in sheet_names:
        wb.create_sheet(title=name)
    return wb

def test_single_exact_pattern_match():
    """Test pattern match 1 element"""
    wb = create_workbook_with_sheets(["fOut_test1", "data_sheet", "summary"])
    matched = extract_fout_sheets(wb, [r"^fOut_"])
    assert matched == ["fOut_test1"]

def test_multiple_patterns_match():
    """Test multiple patterns"""
    wb = create_workbook_with_sheets(
        [
            "fOut_test1",
            "data_sheet",
            "fOut_data",
            "results",
            "data_export"])
    matched = extract_fout_sheets(wb, [r"^fOut_", r"^data_"])
    assert matched == ["fOut_test1", "data_sheet", "fOut_data", "data_export"]

def test_no_matches_raises_value_error():
    """Test non-matchin sheet names"""
    wb = create_workbook_with_sheets(
        [
            "summary",
            "results",
            "report"
        ])
    with pytest.raises(ValueError, match="No sheets matching patterns"):
        extract_fout_sheets(wb, [r"^fOut_", r"^data_"])

def test_case_sensitive_behavior():
    """Test with case variance in name"""
    wb = create_workbook_with_sheets(["fout_test", "FOUT_data", "fOut_valid"])
    matched = extract_fout_sheets(wb, [r"^fOut_"])
    assert matched == ["fOut_valid"]

def test_partial_match_pattern():
    """Partial matching within string"""
    wb = create_workbook_with_sheets(
        [
            "random_fOut_test",
            "pre_data_sheet",
            "info_fOut_data"
        ])
    # Using a pattern that can match anywhere (re.search instead of re.match is needed for this)
    # but since function uses re.match, these shouldn't match
    with pytest.raises(ValueError):
        extract_fout_sheets(wb, [r"fOut_"])
